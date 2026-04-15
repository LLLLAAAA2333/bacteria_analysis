import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook

from bacteria_analysis.model_space import (
    build_stimulus_sample_map,
    build_model_feature_matrix,
    build_model_rdm,
    build_metabolite_annotation_skeleton,
    load_stimulus_sample_map,
    load_model_registry,
    read_metabolite_matrix,
    resolve_model_inputs,
    _validate_correlation_distance_inputs,
)

MODEL_REGISTRY_COLUMNS = [
    "model_id",
    "model_label",
    "model_tier",
    "model_status",
    "feature_kind",
    "distance_kind",
    "description",
    "authority",
    "notes",
]
MODEL_MEMBERSHIP_COLUMNS = [
    "model_id",
    "metabolite_name",
    "membership_source",
    "review_status",
    "ambiguous_flag",
    "notes",
]
METABOLITE_ANNOTATION_COLUMNS = [
    "metabolite_name",
    "superclass",
    "subclass",
    "pathway_tag",
    "annotation_source",
    "review_status",
    "ambiguous_flag",
    "notes",
]


def _make_task3_resolved_inputs(*, matrix_rows, mapping_rows, registry_rows, membership_rows):
    matrix = pd.DataFrame.from_records(matrix_rows).set_index("sample_id")
    matrix.index = matrix.index.astype(str)
    matrix.index.name = "sample_id"

    mapping = pd.DataFrame.from_records(mapping_rows, columns=["sample_id", "stimulus", "stim_name"])
    mapping["sample_id"] = mapping["sample_id"].astype(str)
    mapping["stimulus"] = mapping["stimulus"].astype(str)
    mapping["stim_name"] = mapping["stim_name"].astype(str)

    registry = pd.DataFrame.from_records(registry_rows, columns=MODEL_REGISTRY_COLUMNS)
    registry["model_id"] = registry["model_id"].astype(str).str.lower()
    registry["model_tier"] = registry["model_tier"].astype(str).str.lower()
    registry["model_status"] = registry["model_status"].astype(str).str.lower()
    registry["feature_kind"] = registry["feature_kind"].astype(str).str.lower()
    registry["distance_kind"] = registry["distance_kind"].astype(str).str.lower()
    registry_resolved = registry.copy()
    registry_resolved["is_primary_family"] = registry_resolved["model_tier"].eq("primary")
    registry_resolved["is_supplementary_family"] = registry_resolved["model_tier"].eq("supplementary")

    membership = pd.DataFrame.from_records(membership_rows, columns=MODEL_MEMBERSHIP_COLUMNS)
    if membership.empty:
        membership = pd.DataFrame(columns=MODEL_MEMBERSHIP_COLUMNS)
    membership["model_id"] = membership["model_id"].astype(str).str.lower()
    membership["metabolite_name"] = membership["metabolite_name"].astype(str)
    membership_resolved = membership.copy()

    annotation = pd.DataFrame.from_records(
        [
            {
                "metabolite_name": metabolite_name,
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            }
            for metabolite_name in matrix.columns.astype(str).tolist()
        ],
        columns=METABOLITE_ANNOTATION_COLUMNS,
    )
    return {
        "matrix": matrix,
        "stimulus_sample_map": mapping,
        "metabolite_annotation": annotation,
        "model_registry": registry,
        "model_membership": membership,
        "model_registry_resolved": registry_resolved,
        "model_membership_resolved": membership_resolved,
    }


def _write_stage3_model_space_files(root, *, registry_rows, membership_rows, annotation_rows):
    pd.DataFrame.from_records(registry_rows, columns=MODEL_REGISTRY_COLUMNS).to_csv(
        root / "model_registry.csv", index=False
    )
    pd.DataFrame.from_records(membership_rows, columns=MODEL_MEMBERSHIP_COLUMNS).to_csv(
        root / "model_membership.csv", index=False
    )
    pd.DataFrame.from_records(annotation_rows, columns=METABOLITE_ANNOTATION_COLUMNS).to_csv(
        root / "metabolite_annotation.csv", index=False
    )


def test_load_stimulus_sample_map_requires_unique_sample_ids(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Stimulus A"},
            {"sample_id": "A001", "stimulus": "stimulus_b", "stim_name": "Stimulus B"},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Stimulus C"},
        ]
    ).to_csv(root / "duplicate_stimulus_sample_map.csv", index=False)

    with pytest.raises(ValueError, match="sample_id values must be unique"):
        load_stimulus_sample_map(root / "duplicate_stimulus_sample_map.csv")


def test_load_stimulus_sample_map_requires_unique_stimulus_values(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Stimulus A"},
            {"sample_id": "A002", "stimulus": "stimulus_a", "stim_name": "Stimulus B"},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Stimulus C"},
        ]
    ).to_csv(root / "duplicate_stimulus_values_map.csv", index=False)

    with pytest.raises(ValueError, match="stimulus values must be unique"):
        load_stimulus_sample_map(root / "duplicate_stimulus_values_map.csv")


def test_load_stimulus_sample_map_allows_duplicate_stim_name_values(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Shared Stimulus"},
            {"sample_id": "A002", "stimulus": "stimulus_b", "stim_name": "Shared Stimulus"},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Shared Stimulus"},
        ]
    ).to_csv(root / "duplicate_stim_name_map.csv", index=False)

    frame = load_stimulus_sample_map(root / "duplicate_stim_name_map.csv")
    assert frame["stim_name"].tolist() == ["Shared Stimulus", "Shared Stimulus", "Shared Stimulus"]


def test_load_stimulus_sample_map_rejects_blank_stim_name_values(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Stimulus A"},
            {"sample_id": "A002", "stimulus": "stimulus_b", "stim_name": ""},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Stimulus C"},
        ]
    ).to_csv(root / "blank_stim_name_map.csv", index=False)

    with pytest.raises(ValueError, match="stim_name values must be non-empty"):
        load_stimulus_sample_map(root / "blank_stim_name_map.csv")


def test_build_stimulus_sample_map_extracts_sample_id_from_stim_name():
    metadata = pd.DataFrame.from_records(
        [
            {"stimulus": "b34_0", "stim_name": "A226 stationary"},
            {"stimulus": "b35_0", "stim_name": "A228 stationary"},
            {"stimulus": "b34_0", "stim_name": "A226 stationary"},
        ]
    )

    mapping = build_stimulus_sample_map(metadata, matrix_sample_ids=pd.Index(["A226", "A228"]))

    assert mapping.to_dict(orient="records") == [
        {"stimulus": "b34_0", "stim_name": "A226 stationary", "sample_id": "A226"},
        {"stimulus": "b35_0", "stim_name": "A228 stationary", "sample_id": "A228"},
    ]


def test_build_stimulus_sample_map_rejects_conflicting_stim_name():
    metadata = pd.DataFrame.from_records(
        [
            {"stimulus": "b34_0", "stim_name": "A226 stationary"},
            {"stimulus": "b34_0", "stim_name": "A227 stationary"},
        ]
    )

    with pytest.raises(ValueError, match="exactly one stim_name"):
        build_stimulus_sample_map(metadata, matrix_sample_ids=pd.Index(["A226", "A227"]))


def test_read_metabolite_matrix_loads_expected_sample_ids(tmp_path):
    matrix_path = tmp_path / "matrix.xlsx"
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "feature_1": 0.1, "feature_2": 1.0},
            {"sample_id": "A002", "feature_1": 0.2, "feature_2": 0.8},
            {"sample_id": "A003", "feature_1": 0.3, "feature_2": 0.6},
        ]
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    matrix = read_metabolite_matrix(matrix_path)
    assert matrix.index.tolist() == ["A001", "A002", "A003"]


def test_read_metabolite_matrix_strips_metabolite_header_whitespace(tmp_path):
    matrix_path = tmp_path / "whitespace_metabolite_headers.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["sample_id", " feature_1 ", "feature_2  "])
    sheet.append(["A001", 0.1, 1.0])
    sheet.append(["A002", 0.2, 0.8])
    workbook.save(matrix_path)

    matrix = read_metabolite_matrix(matrix_path)

    assert matrix.columns.tolist() == ["feature_1", "feature_2"]


def test_read_metabolite_matrix_canonicalizes_known_metabolite_headers(tmp_path):
    matrix_path = tmp_path / "canonical_metabolite_headers.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "sample_id",
            "Beta-Muricholic acid (\u03b2-MCA)",
            "Adenosine-5\u2032-triphosphate(ATP)",
            "Tauro-\u03b1-muricholic acid (\u03c9-TMCA)",
        ]
    )
    sheet.append(["A001", 0.1, 1.0, 2.0])
    sheet.append(["A002", 0.2, 0.8, 1.5])
    workbook.save(matrix_path)

    matrix = read_metabolite_matrix(matrix_path)

    assert matrix.columns.tolist() == [
        "Beta-Muricholic acid (beta-MCA)",
        "Adenosine-5'-triphosphate(ATP)",
        "Tauro-omega-muricholic acid (omega-TMCA)",
    ]


def test_read_metabolite_matrix_rejects_blank_sample_id_cells(tmp_path):
    matrix_path = tmp_path / "blank_matrix.xlsx"
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "feature_1": 0.1, "feature_2": 1.0},
            {"sample_id": None, "feature_1": 0.2, "feature_2": 0.8},
            {"sample_id": "A003", "feature_1": 0.3, "feature_2": 0.6},
        ]
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    with pytest.raises(ValueError, match="sample_id values must be non-empty"):
        read_metabolite_matrix(matrix_path)


def test_read_metabolite_matrix_rejects_duplicate_metabolite_columns(tmp_path):
    matrix_path = tmp_path / "duplicate_metabolite_columns.xlsx"
    pd.DataFrame(
        [
            ["A001", 0.1, 1.0],
            ["A002", 0.2, 0.8],
            ["A003", 0.3, 0.6],
        ],
        columns=["sample_id", "feature_1", "feature_1"],
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    with pytest.raises(ValueError, match="metabolite column names must be unique"):
        read_metabolite_matrix(matrix_path)


def test_read_metabolite_matrix_rejects_blank_metabolite_headers(tmp_path):
    matrix_path = tmp_path / "blank_metabolite_header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["sample_id", None, "feature_2"])
    sheet.append(["A001", 0.1, 1.0])
    sheet.append(["A002", 0.2, 0.8])
    workbook.save(matrix_path)

    with pytest.raises(ValueError, match="metabolite column names must be non-empty"):
        read_metabolite_matrix(matrix_path)


def test_read_metabolite_matrix_uses_first_worksheet_for_validation_and_loading(tmp_path):
    matrix_path = tmp_path / "multi_sheet_matrix.xlsx"
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = "matrix"
    first_sheet.append(["sample_id", "feature_1", "feature_2"])
    first_sheet.append(["A001", 0.1, 1.0])
    first_sheet.append(["A002", 0.2, 0.8])
    second_sheet = workbook.create_sheet("active_sheet")
    second_sheet.append(["sample_id", None, "feature_2"])
    second_sheet.append(["B001", 0.4, 0.6])
    workbook.active = 1
    workbook.save(matrix_path)

    matrix = read_metabolite_matrix(matrix_path)

    assert matrix.index.tolist() == ["A001", "A002"]
    assert matrix.columns.tolist() == ["feature_1", "feature_2"]


@pytest.mark.parametrize(
    ("annotation_rows", "message"),
    [
        (
            [
                {
                    "metabolite_name": "feature_1",
                    "superclass": "",
                    "subclass": "",
                    "pathway_tag": "",
                    "annotation_source": "",
                    "review_status": "",
                    "ambiguous_flag": False,
                    "notes": "",
                }
            ],
            "annotation metabolites must cover all matrix metabolites: feature_2",
        ),
        (
            [
                {
                    "metabolite_name": "feature_1",
                    "superclass": "",
                    "subclass": "",
                    "pathway_tag": "",
                    "annotation_source": "",
                    "review_status": "",
                    "ambiguous_flag": False,
                    "notes": "",
                },
                {
                    "metabolite_name": "feature_2",
                    "superclass": "",
                    "subclass": "",
                    "pathway_tag": "",
                    "annotation_source": "",
                    "review_status": "",
                    "ambiguous_flag": False,
                    "notes": "",
                },
                {
                    "metabolite_name": "feature_3",
                    "superclass": "",
                    "subclass": "",
                    "pathway_tag": "",
                    "annotation_source": "",
                    "review_status": "",
                    "ambiguous_flag": False,
                    "notes": "",
                },
            ],
            "annotation metabolites must exist in the matrix: feature_3",
        ),
    ],
)
def test_resolve_model_inputs_rejects_annotation_mismatches(tmp_path, annotation_rows, message):
    root = tmp_path / "model_space"
    root.mkdir()
    matrix_path = root / "matrix.xlsx"
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "feature_1": 0.1, "feature_2": 1.0},
            {"sample_id": "A002", "feature_1": 0.2, "feature_2": 0.8},
            {"sample_id": "A003", "feature_1": 0.3, "feature_2": 0.6},
        ]
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Stimulus A"},
            {"sample_id": "A002", "stimulus": "stimulus_b", "stim_name": "Stimulus B"},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Stimulus C"},
        ]
    ).to_csv(root / "stimulus_sample_map.csv", index=False)

    _write_stage3_model_space_files(
        root,
        registry_rows=[
            {
                "model_id": "global_profile",
                "model_label": "Global Metabolite Profile",
                "model_tier": "primary",
                "model_status": "primary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "All matrix metabolites",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[],
        annotation_rows=annotation_rows,
    )

    with pytest.raises(ValueError, match=message):
        resolve_model_inputs(root, matrix_path)


def test_resolve_model_inputs_accepts_broad_union_text_when_registry_metadata_is_valid(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    matrix_path = root / "matrix.xlsx"
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "feature_1": 0.1, "feature_2": 1.0},
            {"sample_id": "A002", "feature_1": 0.2, "feature_2": 0.8},
            {"sample_id": "A003", "feature_1": 0.3, "feature_2": 0.6},
        ]
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Stimulus A"},
            {"sample_id": "A002", "stimulus": "stimulus_b", "stim_name": "Stimulus B"},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Stimulus C"},
        ]
    ).to_csv(root / "stimulus_sample_map.csv", index=False)

    _write_stage3_model_space_files(
        root,
        registry_rows=[
            {
                "model_id": "global_profile",
                "model_label": "Global Metabolite Profile",
                "model_tier": "primary",
                "model_status": "primary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "All matrix metabolites",
                "authority": "user",
                "notes": "",
            },
            {
                "model_id": "broad_union",
                "model_label": "Broad Union Model",
                "model_tier": "primary",
                "model_status": "supplementary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "Exploratory broad union",
                "authority": "exploratory",
                "notes": "",
            },
        ],
        membership_rows=[
            {
                "model_id": "global_profile",
                "metabolite_name": "feature_1",
                "membership_source": "seed",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            }
        ],
        annotation_rows=[
            {
                "metabolite_name": "feature_1",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "metabolite_name": "feature_2",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    resolved = resolve_model_inputs(root, matrix_path)
    broad_union_row = resolved["model_registry_resolved"].loc[
        resolved["model_registry_resolved"]["model_id"] == "broad_union"
    ].iloc[0]

    assert broad_union_row["model_tier"] == "primary"
    assert broad_union_row["model_status"] == "supplementary"


def test_load_model_registry_rejects_case_variant_duplicate_model_ids(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    pd.DataFrame.from_records(
        [
            {
                "model_id": "GLOBAL_PROFILE",
                "model_label": "Global Metabolite Profile",
                "model_tier": "primary",
                "model_status": "primary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "All matrix metabolites",
                "authority": "user",
                "notes": "",
            },
            {
                "model_id": "global_profile",
                "model_label": "Global Metabolite Profile",
                "model_tier": "primary",
                "model_status": "primary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "All matrix metabolites",
                "authority": "user",
                "notes": "",
            },
        ],
        columns=MODEL_REGISTRY_COLUMNS,
    ).to_csv(root / "model_registry.csv", index=False)

    with pytest.raises(ValueError, match="model_id values must be unique"):
        load_model_registry(root / "model_registry.csv")


def test_resolve_model_inputs_canonicalizes_case_variant_global_profile_rows(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    matrix_path = root / "matrix.xlsx"
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "feature_1": 0.1, "feature_2": 1.0},
            {"sample_id": "A002", "feature_1": 0.2, "feature_2": 0.8},
            {"sample_id": "A003", "feature_1": 0.3, "feature_2": 0.6},
        ]
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Stimulus A"},
            {"sample_id": "A002", "stimulus": "stimulus_b", "stim_name": "Stimulus B"},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Stimulus C"},
        ]
    ).to_csv(root / "stimulus_sample_map.csv", index=False)

    _write_stage3_model_space_files(
        root,
        registry_rows=[
            {
                "model_id": "GLOBAL_PROFILE",
                "model_label": "Global Metabolite Profile",
                "model_tier": "primary",
                "model_status": "primary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "All matrix metabolites",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[
            {
                "model_id": "GLOBAL_PROFILE",
                "metabolite_name": "feature_1",
                "membership_source": "seed",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            }
        ],
        annotation_rows=[
            {
                "metabolite_name": "feature_1",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "metabolite_name": "feature_2",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    resolved = resolve_model_inputs(root, matrix_path)

    assert resolved["model_registry_resolved"]["model_id"].tolist() == ["global_profile"]
    assert resolved["model_membership_resolved"]["model_id"].tolist() == ["global_profile", "global_profile"]
    assert set(resolved["model_membership_resolved"]["metabolite_name"]) == {"feature_1", "feature_2"}


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("model_tier", "secondary"),
        ("model_status", "pending"),
        ("feature_kind", "rank_abundance"),
        ("distance_kind", "cosine"),
    ],
)
def test_load_model_registry_rejects_invalid_classification_values(tmp_path, field, value):
    root = tmp_path / "model_space"
    root.mkdir()
    registry_rows = [
        {
            "model_id": "global_profile",
            "model_label": "Global Metabolite Profile",
            "model_tier": "primary",
            "model_status": "primary",
            "feature_kind": "continuous_abundance",
            "distance_kind": "correlation",
            "description": "All matrix metabolites",
            "authority": "user",
            "notes": "",
        },
        {
            "model_id": "broad_union",
            "model_label": "Broad Union Model",
            "model_tier": "primary",
            "model_status": "supplementary",
            "feature_kind": "continuous_abundance",
            "distance_kind": "correlation",
            "description": "Exploratory broad union",
            "authority": "exploratory",
            "notes": "",
        },
    ]
    registry_rows[1][field] = value
    pd.DataFrame.from_records(registry_rows, columns=MODEL_REGISTRY_COLUMNS).to_csv(
        root / "model_registry.csv", index=False
    )

    with pytest.raises(ValueError, match=f"{field} values must be one of"):
        load_model_registry(root / "model_registry.csv")


@pytest.mark.parametrize("blank_field", ["model_tier", "model_status"])
def test_load_model_registry_rejects_blank_required_fields(tmp_path, blank_field):
    root = tmp_path / "model_space"
    root.mkdir()
    registry_rows = [
        {
            "model_id": "global_profile",
            "model_label": "Global Metabolite Profile",
            "model_tier": "primary",
            "model_status": "primary",
            "feature_kind": "continuous_abundance",
            "distance_kind": "correlation",
            "description": "All matrix metabolites",
            "authority": "user",
            "notes": "",
        },
        {
            "model_id": "broad_union",
            "model_label": "Broad Union Model",
            "model_tier": "primary",
            "model_status": "supplementary",
            "feature_kind": "continuous_abundance",
            "distance_kind": "correlation",
            "description": "Exploratory broad union",
            "authority": "exploratory",
            "notes": "",
        },
    ]
    registry_rows[1][blank_field] = ""
    pd.DataFrame.from_records(registry_rows, columns=MODEL_REGISTRY_COLUMNS).to_csv(
        root / "model_registry.csv", index=False
    )

    with pytest.raises(ValueError, match=f"{blank_field} values must be non-empty"):
        load_model_registry(root / "model_registry.csv")


def test_load_model_registry_canonicalizes_allowed_value_fields(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    registry_rows = [
        {
            "model_id": "global_profile",
            "model_label": "Global Metabolite Profile",
            "model_tier": "PRIMARY",
            "model_status": "Draft",
            "feature_kind": "Continuous_Abundance",
            "distance_kind": "Euclidean",
            "description": "All matrix metabolites",
            "authority": "user",
            "notes": "",
        },
        {
            "model_id": "broad_union",
            "model_label": "Broad Union Model",
            "model_tier": "Supplementary",
            "model_status": "EXCLUDED",
            "feature_kind": "Binary_Presence",
            "distance_kind": "JACCARD",
            "description": "Exploratory broad union",
            "authority": "exploratory",
            "notes": "",
        },
    ]
    pd.DataFrame.from_records(registry_rows, columns=MODEL_REGISTRY_COLUMNS).to_csv(
        root / "model_registry.csv", index=False
    )

    frame = load_model_registry(root / "model_registry.csv")

    assert frame["model_tier"].tolist() == ["primary", "supplementary"]
    assert frame["model_status"].tolist() == ["draft", "excluded"]
    assert frame["feature_kind"].tolist() == ["continuous_abundance", "binary_presence"]
    assert frame["distance_kind"].tolist() == ["euclidean", "jaccard"]


def test_resolve_model_inputs_accepts_draft_broad_union_text_without_keyword_inference(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    matrix_path = root / "matrix.xlsx"
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "feature_1": 0.1, "feature_2": 1.0},
            {"sample_id": "A002", "feature_1": 0.2, "feature_2": 0.8},
            {"sample_id": "A003", "feature_1": 0.3, "feature_2": 0.6},
        ]
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Stimulus A"},
            {"sample_id": "A002", "stimulus": "stimulus_b", "stim_name": "Stimulus B"},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Stimulus C"},
        ]
    ).to_csv(root / "stimulus_sample_map.csv", index=False)

    _write_stage3_model_space_files(
        root,
        registry_rows=[
            {
                "model_id": "draft_broad_union",
                "model_label": "Draft Broad Union Model",
                "model_tier": "primary",
                "model_status": "draft",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "Draft broad union candidate",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[],
        annotation_rows=[
            {
                "metabolite_name": "feature_1",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "metabolite_name": "feature_2",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    resolved = resolve_model_inputs(root, matrix_path)
    draft_row = resolved["model_registry_resolved"].loc[
        resolved["model_registry_resolved"]["model_id"] == "draft_broad_union"
    ].iloc[0]

    assert draft_row["model_tier"] == "primary"
    assert draft_row["model_status"] == "draft"


def test_resolve_model_inputs_rejects_unknown_membership_metabolite(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    matrix_path = root / "matrix.xlsx"
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "feature_1": 0.1, "feature_2": 1.0},
            {"sample_id": "A002", "feature_1": 0.2, "feature_2": 0.8},
            {"sample_id": "A003", "feature_1": 0.3, "feature_2": 0.6},
        ]
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Stimulus A"},
            {"sample_id": "A002", "stimulus": "stimulus_b", "stim_name": "Stimulus B"},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Stimulus C"},
        ]
    ).to_csv(root / "stimulus_sample_map.csv", index=False)

    _write_stage3_model_space_files(
        root,
        registry_rows=[
            {
                "model_id": "global_profile",
                "model_label": "Global Metabolite Profile",
                "model_tier": "primary",
                "model_status": "primary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "All matrix metabolites",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[
            {
                "model_id": "global_profile",
                "metabolite_name": "feature_typo",
                "membership_source": "seed",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            }
        ],
        annotation_rows=[
            {
                "metabolite_name": "feature_1",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "metabolite_name": "feature_2",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    with pytest.raises(ValueError, match="model_membership metabolites must exist in the matrix: feature_typo"):
        resolve_model_inputs(root, matrix_path)


def test_resolve_model_inputs_accepts_canonical_annotation_names_for_known_matrix_aliases(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    matrix_path = root / "matrix.xlsx"
    pd.DataFrame.from_records(
        [
            {
                "sample_id": "A001",
                "Beta-Muricholic acid (\u03b2-MCA)": 0.1,
                "Adenosine-5\u2032-triphosphate(ATP)": 1.0,
            },
            {
                "sample_id": "A002",
                "Beta-Muricholic acid (\u03b2-MCA)": 0.2,
                "Adenosine-5\u2032-triphosphate(ATP)": 0.8,
            },
            {
                "sample_id": "A003",
                "Beta-Muricholic acid (\u03b2-MCA)": 0.3,
                "Adenosine-5\u2032-triphosphate(ATP)": 0.6,
            },
        ]
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Stimulus A"},
            {"sample_id": "A002", "stimulus": "stimulus_b", "stim_name": "Stimulus B"},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Stimulus C"},
        ]
    ).to_csv(root / "stimulus_sample_map.csv", index=False)

    _write_stage3_model_space_files(
        root,
        registry_rows=[
            {
                "model_id": "global_profile",
                "model_label": "Global Metabolite Profile",
                "model_tier": "primary",
                "model_status": "primary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "All matrix metabolites",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[],
        annotation_rows=[
            {
                "metabolite_name": "Beta-Muricholic acid (beta-MCA)",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "metabolite_name": "Adenosine-5'-triphosphate(ATP)",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    resolved = resolve_model_inputs(root, matrix_path)

    assert resolved["matrix"].columns.tolist() == [
        "Beta-Muricholic acid (beta-MCA)",
        "Adenosine-5'-triphosphate(ATP)",
    ]
    assert set(resolved["metabolite_annotation"]["metabolite_name"]) == {
        "Beta-Muricholic acid (beta-MCA)",
        "Adenosine-5'-triphosphate(ATP)",
    }


def test_resolve_model_inputs_accepts_canonical_membership_names_for_known_matrix_aliases(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    matrix_path = root / "matrix.xlsx"
    pd.DataFrame.from_records(
        [
            {
                "sample_id": "A001",
                "Tauro-\u03b1-muricholic acid (\u03c9-TMCA)": 0.1,
                "Adenosine-5\u2032-triphosphate(ATP)": 1.0,
            },
            {
                "sample_id": "A002",
                "Tauro-\u03b1-muricholic acid (\u03c9-TMCA)": 0.2,
                "Adenosine-5\u2032-triphosphate(ATP)": 0.8,
            },
            {
                "sample_id": "A003",
                "Tauro-\u03b1-muricholic acid (\u03c9-TMCA)": 0.3,
                "Adenosine-5\u2032-triphosphate(ATP)": 0.6,
            },
        ]
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Stimulus A"},
            {"sample_id": "A002", "stimulus": "stimulus_b", "stim_name": "Stimulus B"},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Stimulus C"},
        ]
    ).to_csv(root / "stimulus_sample_map.csv", index=False)

    _write_stage3_model_space_files(
        root,
        registry_rows=[
            {
                "model_id": "bile_acid",
                "model_label": "Bile Acid",
                "model_tier": "primary",
                "model_status": "primary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "Narrow bile acid family",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[
            {
                "model_id": "bile_acid",
                "metabolite_name": "Tauro-omega-muricholic acid (omega-TMCA)",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "model_id": "bile_acid",
                "metabolite_name": "Adenosine-5'-triphosphate(ATP)",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
        annotation_rows=[
            {
                "metabolite_name": "Tauro-omega-muricholic acid (omega-TMCA)",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "metabolite_name": "Adenosine-5'-triphosphate(ATP)",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    resolved = resolve_model_inputs(root, matrix_path)
    feature_matrix, _ = build_model_feature_matrix(resolved, model_id="bile_acid")

    assert feature_matrix.columns.tolist() == [
        "Tauro-omega-muricholic acid (omega-TMCA)",
        "Adenosine-5'-triphosphate(ATP)",
    ]


def test_resolve_model_inputs_allows_primary_supplementary_status_when_not_broad_union(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    matrix_path = root / "matrix.xlsx"
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "feature_1": 0.1, "feature_2": 1.0},
            {"sample_id": "A002", "feature_1": 0.2, "feature_2": 0.8},
            {"sample_id": "A003", "feature_1": 0.3, "feature_2": 0.6},
        ]
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Stimulus A"},
            {"sample_id": "A002", "stimulus": "stimulus_b", "stim_name": "Stimulus B"},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Stimulus C"},
        ]
    ).to_csv(root / "stimulus_sample_map.csv", index=False)

    _write_stage3_model_space_files(
        root,
        registry_rows=[
            {
                "model_id": "bile_acid",
                "model_label": "Bile Acid",
                "model_tier": "primary",
                "model_status": "supplementary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "Narrow bile acid family",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[],
        annotation_rows=[
            {
                "metabolite_name": "feature_1",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "metabolite_name": "feature_2",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    resolved = resolve_model_inputs(root, matrix_path)
    row = resolved["model_registry_resolved"].loc[resolved["model_registry_resolved"]["model_id"] == "bile_acid"].iloc[0]
    assert row["model_status"] == "supplementary"
    assert row["is_primary_family"]


def test_resolve_model_inputs_seeds_global_profile_for_header_only_empty_membership(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    matrix_path = root / "matrix.xlsx"
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "feature_1": 0.1, "feature_2": 1.0},
            {"sample_id": "A002", "feature_1": 0.2, "feature_2": 0.8},
            {"sample_id": "A003", "feature_1": 0.3, "feature_2": 0.6},
        ]
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Stimulus A"},
            {"sample_id": "A002", "stimulus": "stimulus_b", "stim_name": "Stimulus B"},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Stimulus C"},
        ]
    ).to_csv(root / "stimulus_sample_map.csv", index=False)

    _write_stage3_model_space_files(
        root,
        registry_rows=[
            {
                "model_id": "bile_acid",
                "model_label": "Bile Acid",
                "model_tier": "primary",
                "model_status": "draft",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "Narrow bile acid family",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[],
        annotation_rows=[
            {
                "metabolite_name": "feature_1",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "metabolite_name": "feature_2",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    resolved = resolve_model_inputs(root, matrix_path)

    global_profile_rows = resolved["model_membership_resolved"].loc[
        resolved["model_membership_resolved"]["model_id"] == "global_profile"
    ]
    assert set(global_profile_rows["metabolite_name"]) == {"feature_1", "feature_2"}


def test_resolve_model_inputs_backfills_missing_global_profile_metabolites(tmp_path):
    root = tmp_path / "model_space"
    root.mkdir()
    matrix_path = root / "matrix.xlsx"
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "feature_1": 0.1, "feature_2": 1.0},
            {"sample_id": "A002", "feature_1": 0.2, "feature_2": 0.8},
            {"sample_id": "A003", "feature_1": 0.3, "feature_2": 0.6},
        ]
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "stimulus": "stimulus_a", "stim_name": "Stimulus A"},
            {"sample_id": "A002", "stimulus": "stimulus_b", "stim_name": "Stimulus B"},
            {"sample_id": "A003", "stimulus": "stimulus_c", "stim_name": "Stimulus C"},
        ]
    ).to_csv(root / "stimulus_sample_map.csv", index=False)

    _write_stage3_model_space_files(
        root,
        registry_rows=[
            {
                "model_id": "bile_acid",
                "model_label": "Bile Acid",
                "model_tier": "primary",
                "model_status": "draft",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "Narrow bile acid family",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[
            {
                "model_id": "global_profile",
                "metabolite_name": "feature_1",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            }
        ],
        annotation_rows=[
            {
                "metabolite_name": "feature_1",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "metabolite_name": "feature_2",
                "superclass": "",
                "subclass": "",
                "pathway_tag": "",
                "annotation_source": "",
                "review_status": "",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    resolved = resolve_model_inputs(root, matrix_path)

    global_profile_rows = resolved["model_membership_resolved"].loc[
        resolved["model_membership_resolved"]["model_id"] == "global_profile"
    ]
    assert set(global_profile_rows["metabolite_name"]) == {"feature_1", "feature_2"}
    seeded_row = global_profile_rows.loc[global_profile_rows["metabolite_name"] == "feature_2"].iloc[0]
    assert seeded_row["membership_source"] == "matrix_all_columns"


def test_build_annotation_skeleton_emits_all_matrix_metabolites(tmp_path):
    matrix_path = tmp_path / "matrix.xlsx"
    pd.DataFrame.from_records(
        [
            {"sample_id": "A001", "Cholic acid (CA)": 0.10, "Palmitic acid": 0.30},
            {"sample_id": "A002", "Cholic acid (CA)": 0.20, "Palmitic acid": 0.40},
            {"sample_id": "A003", "Cholic acid (CA)": 0.30, "Palmitic acid": 0.50},
        ]
    ).to_excel(matrix_path, index=False, engine="openpyxl")

    annotation = build_metabolite_annotation_skeleton(matrix_path)
    assert {"metabolite_name", "review_status", "ambiguous_flag"}.issubset(annotation.columns)
    assert "Cholic acid (CA)" in set(annotation["metabolite_name"])


def test_build_model_feature_matrix_drops_zero_variance_columns_for_global_profile():
    resolved_inputs = _make_task3_resolved_inputs(
        matrix_rows=[
            {"sample_id": "A001", "keep_1": 0.0, "drop_me": 1.0, "keep_2": 1.0},
            {"sample_id": "A002", "keep_1": 2.0, "drop_me": 1.0, "keep_2": 3.0},
            {"sample_id": "A003", "keep_1": 4.0, "drop_me": 1.0, "keep_2": 5.0},
        ],
        mapping_rows=[
            {"sample_id": "A002", "stimulus": "b2_1", "stim_name": "Stimulus B2"},
            {"sample_id": "A001", "stimulus": "b1_1", "stim_name": "Stimulus B1"},
            {"sample_id": "A003", "stimulus": "b3_1", "stim_name": "Stimulus B3"},
        ],
        registry_rows=[
            {
                "model_id": "global_profile",
                "model_label": "Global Metabolite Profile",
                "model_tier": "primary",
                "model_status": "primary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "All matrix metabolites",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[],
    )

    feature_matrix, qc = build_model_feature_matrix(resolved_inputs, model_id="global_profile")

    assert feature_matrix.index.tolist() == ["b2_1", "b1_1", "b3_1"]
    assert feature_matrix.columns.tolist() == ["keep_1", "keep_2"]
    assert not bool(qc.loc[qc["metabolite_name"] == "drop_me", "retained"].iloc[0])
    assert qc.loc[qc["metabolite_name"] == "drop_me", "filter_reason"].iloc[0] == "zero_variance"
    np.testing.assert_allclose(feature_matrix.mean(axis=0).to_numpy(), np.zeros(2), atol=1e-9)
    assert "model_feature_qc" in resolved_inputs


def test_build_model_rdm_keeps_fixed_stimulus_order():
    resolved_inputs = _make_task3_resolved_inputs(
        matrix_rows=[
            {"sample_id": "A001", "m1": 0.0, "m2": 1.0},
            {"sample_id": "A002", "m1": 1.0, "m2": 0.0},
            {"sample_id": "A003", "m1": 2.0, "m2": 2.0},
        ],
        mapping_rows=[
            {"sample_id": "A003", "stimulus": "b3_1", "stim_name": "Stimulus B3"},
            {"sample_id": "A001", "stimulus": "b1_1", "stim_name": "Stimulus B1"},
            {"sample_id": "A002", "stimulus": "b2_1", "stim_name": "Stimulus B2"},
        ],
        registry_rows=[
            {
                "model_id": "bile_acid",
                "model_label": "Bile Acid",
                "model_tier": "primary",
                "model_status": "primary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "euclidean",
                "description": "Subset model",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[
            {
                "model_id": "bile_acid",
                "metabolite_name": "m1",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "model_id": "bile_acid",
                "metabolite_name": "m2",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    matrix_frame = build_model_rdm(resolved_inputs, model_id="bile_acid")

    assert matrix_frame["stimulus_row"].tolist() == ["b3_1", "b1_1", "b2_1"]
    assert matrix_frame.columns.tolist() == ["stimulus_row", "b3_1", "b1_1", "b2_1"]


def test_build_model_rdm_uses_binary_presence_with_jaccard_distance():
    resolved_inputs = _make_task3_resolved_inputs(
        matrix_rows=[
            {"sample_id": "A001", "p1": 1.0, "p2": 0.0, "p3": 0.0},
            {"sample_id": "A002", "p1": 1.0, "p2": 1.0, "p3": 0.0},
            {"sample_id": "A003", "p1": 0.0, "p2": 1.0, "p3": 1.0},
        ],
        mapping_rows=[
            {"sample_id": "A001", "stimulus": "b1_1", "stim_name": "Stimulus B1"},
            {"sample_id": "A002", "stimulus": "b2_1", "stim_name": "Stimulus B2"},
            {"sample_id": "A003", "stimulus": "b3_1", "stim_name": "Stimulus B3"},
        ],
        registry_rows=[
            {
                "model_id": "presence_profile",
                "model_label": "Presence Profile",
                "model_tier": "supplementary",
                "model_status": "draft",
                "feature_kind": "binary_presence",
                "distance_kind": "jaccard",
                "description": "Binary presence subset",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[
            {
                "model_id": "presence_profile",
                "metabolite_name": "p1",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "model_id": "presence_profile",
                "metabolite_name": "p2",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "model_id": "presence_profile",
                "metabolite_name": "p3",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    matrix_frame = build_model_rdm(resolved_inputs, model_id="presence_profile")

    assert matrix_frame.loc[0, "b2_1"] == pytest.approx(0.5)
    assert matrix_frame.loc[0, "b3_1"] == pytest.approx(1.0)
    assert matrix_frame.loc[1, "b3_1"] == pytest.approx(2.0 / 3.0)
    assert set(resolved_inputs["model_feature_qc"]["threshold"]) == {0.0}


def test_build_model_rdm_rejects_correlation_models_with_no_resolved_features():
    resolved_inputs = _make_task3_resolved_inputs(
        matrix_rows=[
            {"sample_id": "A001", "m1": 0.0, "m2": 1.0},
            {"sample_id": "A002", "m1": 1.0, "m2": 0.0},
            {"sample_id": "A003", "m1": 2.0, "m2": 2.0},
        ],
        mapping_rows=[
            {"sample_id": "A001", "stimulus": "b1_1", "stim_name": "Stimulus B1"},
            {"sample_id": "A002", "stimulus": "b2_1", "stim_name": "Stimulus B2"},
            {"sample_id": "A003", "stimulus": "b3_1", "stim_name": "Stimulus B3"},
        ],
        registry_rows=[
            {
                "model_id": "subset_model",
                "model_label": "Subset Model",
                "model_tier": "supplementary",
                "model_status": "draft",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "Model with no resolved metabolites",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[],
    )

    with pytest.raises(ValueError, match="at least 2 retained features"):
        build_model_rdm(resolved_inputs, model_id="subset_model")


def test_build_model_rdm_rejects_non_correlation_models_with_no_resolved_features():
    resolved_inputs = _make_task3_resolved_inputs(
        matrix_rows=[
            {"sample_id": "A001", "p1": 1.0, "p2": 0.0},
            {"sample_id": "A002", "p1": 0.0, "p2": 1.0},
            {"sample_id": "A003", "p1": 1.0, "p2": 1.0},
        ],
        mapping_rows=[
            {"sample_id": "A001", "stimulus": "b1_1", "stim_name": "Stimulus B1"},
            {"sample_id": "A002", "stimulus": "b2_1", "stim_name": "Stimulus B2"},
            {"sample_id": "A003", "stimulus": "b3_1", "stim_name": "Stimulus B3"},
        ],
        registry_rows=[
            {
                "model_id": "empty_binary",
                "model_label": "Empty Binary",
                "model_tier": "supplementary",
                "model_status": "draft",
                "feature_kind": "binary_presence",
                "distance_kind": "jaccard",
                "description": "Binary model with no resolved metabolites",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[],
    )

    with pytest.raises(ValueError, match="at least 1 retained feature"):
        build_model_rdm(resolved_inputs, model_id="empty_binary")


@pytest.mark.parametrize("bad_value", [np.nan, np.inf, -np.inf])
def test_build_model_rdm_rejects_non_correlation_models_with_non_finite_features(bad_value):
    resolved_inputs = _make_task3_resolved_inputs(
        matrix_rows=[
            {"sample_id": "A001", "p1": 1.0, "p2": 0.0},
            {"sample_id": "A002", "p1": bad_value, "p2": 1.0},
            {"sample_id": "A003", "p1": 1.0, "p2": 1.0},
        ],
        mapping_rows=[
            {"sample_id": "A001", "stimulus": "b1_1", "stim_name": "Stimulus B1"},
            {"sample_id": "A002", "stimulus": "b2_1", "stim_name": "Stimulus B2"},
            {"sample_id": "A003", "stimulus": "b3_1", "stim_name": "Stimulus B3"},
        ],
        registry_rows=[
            {
                "model_id": "non_finite_binary",
                "model_label": "Non-finite Binary",
                "model_tier": "supplementary",
                "model_status": "draft",
                "feature_kind": "binary_presence",
                "distance_kind": "jaccard",
                "description": "Binary model with non-finite feature values",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[
            {
                "model_id": "non_finite_binary",
                "metabolite_name": "p1",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "model_id": "non_finite_binary",
                "metabolite_name": "p2",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    with pytest.raises(ValueError, match="finite feature values"):
        build_model_rdm(resolved_inputs, model_id="non_finite_binary")


def test_build_model_rdm_rejects_non_correlation_models_with_non_finite_log1p_output():
    resolved_inputs = _make_task3_resolved_inputs(
        matrix_rows=[
            {"sample_id": "A001", "m1": 0.0, "m2": 1.0},
            {"sample_id": "A002", "m1": -2.0, "m2": 2.0},
            {"sample_id": "A003", "m1": 1.0, "m2": 3.0},
        ],
        mapping_rows=[
            {"sample_id": "A001", "stimulus": "b1_1", "stim_name": "Stimulus B1"},
            {"sample_id": "A002", "stimulus": "b2_1", "stim_name": "Stimulus B2"},
            {"sample_id": "A003", "stimulus": "b3_1", "stim_name": "Stimulus B3"},
        ],
        registry_rows=[
            {
                "model_id": "non_finite_euclidean",
                "model_label": "Non-finite Euclidean",
                "model_tier": "supplementary",
                "model_status": "draft",
                "feature_kind": "continuous_abundance",
                "distance_kind": "euclidean",
                "description": "Continuous model with invalid log1p output",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[
            {
                "model_id": "non_finite_euclidean",
                "metabolite_name": "m1",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "model_id": "non_finite_euclidean",
                "metabolite_name": "m2",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    with pytest.raises(ValueError, match="values greater than -1"):
        build_model_rdm(resolved_inputs, model_id="non_finite_euclidean")


def test_build_model_rdm_rejects_degenerate_correlation_rows_after_preprocessing():
    resolved_inputs = _make_task3_resolved_inputs(
        matrix_rows=[
            {"sample_id": "A001", "m1": 0.0, "m2": 1.0},
            {"sample_id": "A002", "m1": 1.0, "m2": 0.0},
            {"sample_id": "A003", "m1": 2.0, "m2": 2.0},
        ],
        mapping_rows=[
            {"sample_id": "A001", "stimulus": "b1_1", "stim_name": "Stimulus B1"},
            {"sample_id": "A002", "stimulus": "b2_1", "stim_name": "Stimulus B2"},
            {"sample_id": "A003", "stimulus": "b3_1", "stim_name": "Stimulus B3"},
        ],
        registry_rows=[
            {
                "model_id": "degenerate_correlation",
                "model_label": "Degenerate Correlation",
                "model_tier": "supplementary",
                "model_status": "draft",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "Model with a zero-norm stimulus row after preprocessing",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[
            {
                "model_id": "degenerate_correlation",
                "metabolite_name": "m1",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "model_id": "degenerate_correlation",
                "metabolite_name": "m2",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    with pytest.raises(ValueError, match="non-constant stimulus rows"):
        build_model_rdm(resolved_inputs, model_id="degenerate_correlation")


@pytest.mark.parametrize("bad_value", [np.nan, np.inf, -np.inf])
def test_validate_correlation_distance_inputs_rejects_non_finite_feature_values(bad_value):
    feature_matrix = pd.DataFrame(
        [
            [0.0, 1.0],
            [1.0, bad_value],
            [2.0, 2.0],
        ],
        index=["b1_1", "b2_1", "b3_1"],
        columns=["m1", "m2"],
    )

    with pytest.raises(ValueError, match="finite feature values"):
        _validate_correlation_distance_inputs(feature_matrix, model_id="non_finite_correlation")


def test_build_model_feature_matrix_marks_tiny_primary_model_excluded_from_ranking():
    resolved_inputs = _make_task3_resolved_inputs(
        matrix_rows=[
            {"sample_id": "A001", "m1": 0.0, "m2": 1.0, "m3": 2.0, "m4": 3.0},
            {"sample_id": "A002", "m1": 1.0, "m2": 2.0, "m3": 3.0, "m4": 4.0},
            {"sample_id": "A003", "m1": 2.0, "m2": 3.0, "m3": 4.0, "m4": 5.0},
        ],
        mapping_rows=[
            {"sample_id": "A001", "stimulus": "b1_1", "stim_name": "Stimulus B1"},
            {"sample_id": "A002", "stimulus": "b2_1", "stim_name": "Stimulus B2"},
            {"sample_id": "A003", "stimulus": "b3_1", "stim_name": "Stimulus B3"},
        ],
        registry_rows=[
            {
                "model_id": "tiny_primary",
                "model_label": "Tiny Primary",
                "model_tier": "primary",
                "model_status": "primary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "Too few informative features",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[
            {
                "model_id": "tiny_primary",
                "metabolite_name": "m1",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "model_id": "tiny_primary",
                "metabolite_name": "m2",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "model_id": "tiny_primary",
                "metabolite_name": "m3",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "model_id": "tiny_primary",
                "metabolite_name": "m4",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    feature_matrix, _ = build_model_feature_matrix(resolved_inputs, model_id="tiny_primary")

    assert feature_matrix.shape == (3, 4)
    excluded = resolved_inputs["model_registry_resolved"].loc[
        resolved_inputs["model_registry_resolved"]["model_id"] == "tiny_primary",
        "excluded_from_primary_ranking",
    ]
    assert bool(excluded.iloc[0])


def test_build_model_feature_matrix_clears_stale_qc_when_model_loses_all_features():
    resolved_inputs = _make_task3_resolved_inputs(
        matrix_rows=[
            {"sample_id": "A001", "m1": 0.0, "m2": 1.0},
            {"sample_id": "A002", "m1": 1.0, "m2": 2.0},
            {"sample_id": "A003", "m1": 2.0, "m2": 3.0},
        ],
        mapping_rows=[
            {"sample_id": "A001", "stimulus": "b1_1", "stim_name": "Stimulus B1"},
            {"sample_id": "A002", "stimulus": "b2_1", "stim_name": "Stimulus B2"},
            {"sample_id": "A003", "stimulus": "b3_1", "stim_name": "Stimulus B3"},
        ],
        registry_rows=[
            {
                "model_id": "stale_qc_model",
                "model_label": "Stale QC Model",
                "model_tier": "supplementary",
                "model_status": "draft",
                "feature_kind": "continuous_abundance",
                "distance_kind": "euclidean",
                "description": "Model used to verify QC replacement",
                "authority": "user",
                "notes": "",
            }
        ],
        membership_rows=[
            {
                "model_id": "stale_qc_model",
                "metabolite_name": "m1",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
            {
                "model_id": "stale_qc_model",
                "metabolite_name": "m2",
                "membership_source": "manual",
                "review_status": "reviewed",
                "ambiguous_flag": False,
                "notes": "",
            },
        ],
    )

    feature_matrix, feature_qc = build_model_feature_matrix(resolved_inputs, model_id="stale_qc_model")

    assert feature_matrix.shape == (3, 2)
    assert not feature_qc.empty

    resolved_inputs["model_membership_resolved"] = resolved_inputs["model_membership_resolved"].loc[
        resolved_inputs["model_membership_resolved"]["model_id"] != "stale_qc_model"
    ].copy()
    feature_matrix, feature_qc = build_model_feature_matrix(resolved_inputs, model_id="stale_qc_model")

    assert feature_matrix.shape == (3, 0)
    assert feature_qc.empty
    assert resolved_inputs["model_feature_qc"].empty
