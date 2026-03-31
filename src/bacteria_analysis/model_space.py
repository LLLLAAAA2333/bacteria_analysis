"""Loaders for Stage 3 model-space inputs."""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

STIMULUS_SAMPLE_MAP_REQUIRED_COLUMNS = ("stimulus", "stim_name", "sample_id")
METABOLITE_ANNOTATION_REQUIRED_COLUMNS = (
    "metabolite_name",
    "superclass",
    "subclass",
    "pathway_tag",
    "annotation_source",
    "review_status",
    "ambiguous_flag",
    "notes",
)
MODEL_REGISTRY_REQUIRED_COLUMNS = (
    "model_id",
    "model_label",
    "model_tier",
    "model_status",
    "feature_kind",
    "distance_kind",
    "description",
    "authority",
    "notes",
)
MODEL_MEMBERSHIP_REQUIRED_COLUMNS = ("model_id", "metabolite_name")
MODEL_MEMBERSHIP_OPTIONAL_COLUMNS = ("membership_source", "review_status", "ambiguous_flag", "notes")
PRIMARY_TIER_VALUE = "primary"
SUPPLEMENTARY_TIER_VALUE = "supplementary"
MODEL_TIER_ALLOWED_VALUES = (PRIMARY_TIER_VALUE, SUPPLEMENTARY_TIER_VALUE)
MODEL_STATUS_ALLOWED_VALUES = (PRIMARY_TIER_VALUE, SUPPLEMENTARY_TIER_VALUE, "draft", "excluded")
FEATURE_KIND_ALLOWED_VALUES = ("continuous_abundance", "binary_presence")
DISTANCE_KIND_ALLOWED_VALUES = ("correlation", "euclidean", "jaccard")
BOOL_TRUE_VALUES = {"true", "1", "yes", "y", "t"}
BOOL_FALSE_VALUES = {"false", "0", "no", "n", "f", ""}
GLOBAL_PROFILE_MODEL_ID = "global_profile"
PRIMARY_FEATURE_COUNT_MINIMUM = 5
DEFAULT_BINARY_PRESENCE_THRESHOLD = 0.0
MODEL_FEATURE_QC_COLUMNS = (
    "model_id",
    "metabolite_name",
    "feature_kind",
    "retained",
    "status",
    "filter_reason",
    "threshold",
)


def load_stimulus_sample_map(path: str | Path) -> pd.DataFrame:
    frame = pd.read_csv(Path(path), dtype=str).fillna("")
    return _validate_stimulus_sample_map(frame)


def load_metabolite_annotation(path: str | Path) -> pd.DataFrame:
    frame = pd.read_csv(Path(path), dtype=str).fillna("")
    return _validate_metabolite_annotation(frame)


def load_model_registry(path: str | Path) -> pd.DataFrame:
    frame = pd.read_csv(Path(path), dtype=str).fillna("")
    return _validate_model_registry(frame)


def load_model_membership(path: str | Path) -> pd.DataFrame:
    frame = pd.read_csv(Path(path), dtype=str).fillna("")
    return _validate_model_membership(frame)


def read_metabolite_matrix(path: str | Path) -> pd.DataFrame:
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        first_sheet = workbook.worksheets[0]
        header_row = next(first_sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    finally:
        workbook.close()

    header_values = ["" if value is None else str(value).strip() for value in header_row]
    if header_values and pd.Index(header_values).duplicated().any():
        raise ValueError("metabolite column names must be unique")

    frame = pd.read_excel(path, engine="openpyxl", sheet_name=0)
    return _normalize_matrix_frame(frame)


def build_metabolite_annotation_skeleton(matrix_path: str | Path) -> pd.DataFrame:
    matrix = read_metabolite_matrix(matrix_path)
    metabolite_names = matrix.columns.tolist()
    return pd.DataFrame(
        {
            "metabolite_name": metabolite_names,
            "superclass": [""] * len(metabolite_names),
            "subclass": [""] * len(metabolite_names),
            "pathway_tag": [""] * len(metabolite_names),
            "annotation_source": [""] * len(metabolite_names),
            "review_status": [""] * len(metabolite_names),
            "ambiguous_flag": [False] * len(metabolite_names),
            "notes": [""] * len(metabolite_names),
        }
    )


def resolve_model_inputs(model_input_root: str | Path, matrix_path: str | Path) -> dict[str, pd.DataFrame]:
    model_input_root = Path(model_input_root)
    mapping = load_stimulus_sample_map(model_input_root / "stimulus_sample_map.csv")
    annotation = load_metabolite_annotation(model_input_root / "metabolite_annotation.csv")
    registry = load_model_registry(model_input_root / "model_registry.csv")
    membership = load_model_membership(model_input_root / "model_membership.csv")
    return _resolve_stage3_inputs(mapping, annotation, registry, membership, matrix_path)


def build_model_feature_matrix(
    resolved_inputs: dict[str, pd.DataFrame],
    model_id: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    normalized_model_id = str(model_id).strip().lower()
    registry_row = _get_model_registry_row(resolved_inputs, normalized_model_id)
    ordered_matrix = _select_model_input_matrix(resolved_inputs, normalized_model_id)
    feature_kind = str(registry_row["feature_kind"]).strip().lower()

    if feature_kind == "continuous_abundance":
        transformed = np.log1p(ordered_matrix.astype(float))
        filtered, feature_qc = _drop_zero_variance_features(
            transformed,
            model_id=normalized_model_id,
            feature_kind=feature_kind,
        )
        feature_matrix = _zscore_columns(filtered)
    elif feature_kind == "binary_presence":
        transformed = ordered_matrix.astype(float).gt(DEFAULT_BINARY_PRESENCE_THRESHOLD).astype(int)
        filtered, feature_qc = _drop_zero_variance_features(
            transformed,
            model_id=normalized_model_id,
            feature_kind=feature_kind,
            threshold=DEFAULT_BINARY_PRESENCE_THRESHOLD,
        )
        feature_matrix = filtered
    else:
        raise ValueError(f"unsupported feature_kind for model {normalized_model_id!r}: {feature_kind}")

    _update_primary_ranking_exclusion(
        resolved_inputs,
        normalized_model_id,
        retained_feature_count=feature_matrix.shape[1],
    )
    _store_model_feature_qc(resolved_inputs, feature_qc)
    return feature_matrix, feature_qc


def build_model_rdm(resolved_inputs: dict[str, pd.DataFrame], model_id: str) -> pd.DataFrame:
    normalized_model_id = str(model_id).strip().lower()
    feature_matrix, _ = build_model_feature_matrix(resolved_inputs, normalized_model_id)
    registry_row = _get_model_registry_row(resolved_inputs, normalized_model_id)
    distance_kind = str(registry_row["distance_kind"]).strip().lower()

    if distance_kind == "correlation":
        _validate_correlation_distance_inputs(feature_matrix, model_id=normalized_model_id)
    distance_matrix = _compute_model_distance_matrix(feature_matrix, distance_kind=distance_kind)
    matrix = pd.DataFrame(distance_matrix, index=feature_matrix.index, columns=feature_matrix.index)
    frame = matrix.copy()
    frame.insert(0, "stimulus_row", frame.index.astype(str))
    return frame.reset_index(drop=True)


def _validate_stimulus_sample_map(frame: pd.DataFrame) -> pd.DataFrame:
    normalized = _require_columns(frame, STIMULUS_SAMPLE_MAP_REQUIRED_COLUMNS, "stimulus_sample_map")
    for column in STIMULUS_SAMPLE_MAP_REQUIRED_COLUMNS:
        normalized[column] = normalized[column].astype(str).str.strip()

    for column in STIMULUS_SAMPLE_MAP_REQUIRED_COLUMNS:
        _require_non_empty(normalized, column)
    for column in ("stimulus", "sample_id"):
        _require_unique(normalized, column, "stimulus_sample_map")

    return normalized


def _validate_metabolite_annotation(frame: pd.DataFrame) -> pd.DataFrame:
    normalized = _require_columns(frame, METABOLITE_ANNOTATION_REQUIRED_COLUMNS, "metabolite_annotation")
    for column in METABOLITE_ANNOTATION_REQUIRED_COLUMNS:
        if column == "ambiguous_flag":
            continue
        normalized[column] = normalized[column].astype(str).str.strip()

    normalized["ambiguous_flag"] = _coerce_boolean_column(normalized["ambiguous_flag"], "ambiguous_flag")
    _require_non_empty(normalized, "metabolite_name")
    _require_unique(normalized, "metabolite_name", "metabolite_annotation")
    return normalized


def _validate_model_registry(frame: pd.DataFrame) -> pd.DataFrame:
    normalized = _require_columns(frame, MODEL_REGISTRY_REQUIRED_COLUMNS, "model_registry")
    for column in MODEL_REGISTRY_REQUIRED_COLUMNS:
        normalized[column] = normalized[column].astype(str).str.strip()

    normalized["model_id"] = normalized["model_id"].astype(str).str.strip().str.lower()
    for column in ("model_id", "model_tier", "model_status"):
        _require_non_empty(normalized, column)
    _require_allowed_values(normalized, "model_tier", MODEL_TIER_ALLOWED_VALUES)
    _require_allowed_values(normalized, "model_status", MODEL_STATUS_ALLOWED_VALUES)
    _require_allowed_values(normalized, "feature_kind", FEATURE_KIND_ALLOWED_VALUES)
    _require_allowed_values(normalized, "distance_kind", DISTANCE_KIND_ALLOWED_VALUES)
    for column in ("model_tier", "model_status", "feature_kind", "distance_kind"):
        normalized[column] = normalized[column].astype(str).str.strip().str.lower()
    _require_unique(normalized, "model_id", "model_registry")
    return normalized


def _validate_model_membership(frame: pd.DataFrame) -> pd.DataFrame:
    normalized = _require_columns(frame, MODEL_MEMBERSHIP_REQUIRED_COLUMNS, "model_membership")

    for column in MODEL_MEMBERSHIP_REQUIRED_COLUMNS:
        normalized[column] = normalized[column].astype(str).str.strip()
    for column in MODEL_MEMBERSHIP_OPTIONAL_COLUMNS:
        if column not in normalized.columns:
            normalized[column] = ""
        elif column != "ambiguous_flag":
            normalized[column] = normalized[column].astype(str).str.strip()

    normalized["model_id"] = normalized["model_id"].astype(str).str.strip().str.lower()
    normalized["ambiguous_flag"] = _coerce_boolean_column(normalized["ambiguous_flag"], "ambiguous_flag")
    if not normalized.empty:
        _require_non_empty(normalized, "model_id")
        _require_non_empty(normalized, "metabolite_name")
        _require_unique_pairs(normalized, ("model_id", "metabolite_name"), "model_membership")

    return normalized


def _resolve_stage3_inputs(
    mapping: pd.DataFrame,
    annotation: pd.DataFrame,
    registry: pd.DataFrame,
    membership: pd.DataFrame,
    matrix_path: str | Path,
) -> dict[str, pd.DataFrame]:
    matrix = read_metabolite_matrix(matrix_path)
    _validate_mapping_against_matrix(mapping, matrix)
    _validate_annotation_against_matrix(annotation, matrix)
    _validate_membership_against_matrix(membership, matrix)

    resolved_registry = _seed_global_profile_registry(registry)
    resolved_membership = _seed_global_profile_membership(membership, matrix)
    resolved_membership = _validate_membership_against_registry(resolved_membership, resolved_registry)
    resolved_registry["is_primary_family"] = resolved_registry["model_tier"].str.lower().eq(PRIMARY_TIER_VALUE)
    resolved_registry["is_supplementary_family"] = resolved_registry["model_tier"].str.lower().eq(
        SUPPLEMENTARY_TIER_VALUE
    )

    return {
        "matrix": matrix,
        "stimulus_sample_map": mapping.copy(),
        "metabolite_annotation": annotation.copy(),
        "model_registry": registry.copy(),
        "model_membership": membership.copy(),
        "model_registry_resolved": resolved_registry,
        "model_membership_resolved": resolved_membership,
    }


def _normalize_matrix_frame(frame: pd.DataFrame) -> pd.DataFrame:
    normalized = frame.copy()
    if normalized.empty:
        return normalized

    if "sample_id" in normalized.columns:
        sample_column = "sample_id"
    else:
        sample_column = normalized.columns[0]

    sample_ids = normalized[sample_column]
    if sample_ids.isna().any():
        raise ValueError("sample_id values must be non-empty")

    sample_ids = sample_ids.astype(str).str.strip()
    if (sample_ids == "").any():
        raise ValueError("sample_id values must be non-empty")
    if sample_ids.duplicated().any():
        raise ValueError("sample_id values must be unique")

    metabolite_columns = [column for column in normalized.columns if column != sample_column]
    metabolite_column_names = ["" if column is None else str(column).strip() for column in metabolite_columns]
    if any(not name or name.lower().startswith("unnamed:") for name in metabolite_column_names):
        raise ValueError("metabolite column names must be non-empty")
    if pd.Index(metabolite_column_names).duplicated().any():
        raise ValueError("metabolite column names must be unique")

    normalized[sample_column] = sample_ids
    normalized = normalized.set_index(sample_column)
    normalized.index.name = "sample_id"
    return normalized


def _require_columns(frame: pd.DataFrame, required_columns: Iterable[str], label: str) -> pd.DataFrame:
    missing = [column for column in required_columns if column not in frame.columns]
    if missing:
        raise ValueError(f"{label} must include columns: {', '.join(missing)}")
    return frame.copy()


def _require_non_empty(frame: pd.DataFrame, column: str) -> None:
    values = frame[column].astype(str).str.strip()
    if (values == "").any():
        raise ValueError(f"{column} values must be non-empty")


def _require_unique(frame: pd.DataFrame, column: str, label: str) -> None:
    if frame[column].duplicated().any():
        raise ValueError(f"{label} {column} values must be unique")


def _require_allowed_values(frame: pd.DataFrame, column: str, allowed_values: tuple[str, ...]) -> None:
    normalized = frame[column].astype(str).str.strip().str.lower()
    if not normalized.isin(allowed_values).all():
        allowed = ", ".join(allowed_values)
        raise ValueError(f"{column} values must be one of: {allowed}")


def _require_unique_pairs(frame: pd.DataFrame, columns: tuple[str, str], label: str) -> None:
    duplicated = frame.duplicated(list(columns))
    if duplicated.any():
        left, right = columns
        raise ValueError(f"{label} {left} and {right} pairs must be unique")


def _coerce_boolean_column(series: pd.Series, column_name: str) -> pd.Series:
    normalized = series.astype(str).str.strip().str.lower()
    invalid = ~normalized.isin(BOOL_TRUE_VALUES | BOOL_FALSE_VALUES)
    if invalid.any():
        raise ValueError(f"{column_name} values must be boolean-like")
    return normalized.isin(BOOL_TRUE_VALUES)


def _validate_mapping_against_matrix(mapping: pd.DataFrame, matrix: pd.DataFrame) -> None:
    matrix_sample_ids = set(matrix.index.astype(str))
    mapped_sample_ids = set(mapping["sample_id"].astype(str))
    missing = sorted(mapped_sample_ids.difference(matrix_sample_ids))
    if missing:
        raise ValueError(f"mapped sample_id values must exist in the matrix: {', '.join(missing)}")


def _validate_annotation_against_matrix(annotation: pd.DataFrame, matrix: pd.DataFrame) -> None:
    matrix_metabolites = set(matrix.columns.astype(str))
    annotation_metabolites = set(annotation["metabolite_name"].astype(str))
    missing = sorted(matrix_metabolites.difference(annotation_metabolites))
    if missing:
        raise ValueError(f"annotation metabolites must cover all matrix metabolites: {', '.join(missing)}")

    extra = sorted(annotation_metabolites.difference(matrix_metabolites))
    if extra:
        raise ValueError(f"annotation metabolites must exist in the matrix: {', '.join(extra)}")


def _seed_global_profile_membership(membership: pd.DataFrame, matrix: pd.DataFrame) -> pd.DataFrame:
    resolved = _ensure_membership_columns(membership)
    matrix_metabolites = matrix.columns.astype(str).tolist()
    global_profile_mask = resolved["model_id"].astype(str).str.strip().str.lower() == GLOBAL_PROFILE_MODEL_ID
    global_profile_rows = resolved.loc[global_profile_mask]

    seed_model_id = GLOBAL_PROFILE_MODEL_ID
    if not global_profile_rows.empty:
        seed_model_id = global_profile_rows["model_id"].astype(str).iloc[0].strip()

    existing_metabolites = set(global_profile_rows["metabolite_name"].astype(str))
    missing_metabolites = [name for name in matrix_metabolites if name not in existing_metabolites]

    if not missing_metabolites:
        return resolved

    seeded_rows = pd.DataFrame(
        {
            "model_id": [seed_model_id] * len(missing_metabolites),
            "metabolite_name": missing_metabolites,
            "membership_source": ["matrix_all_columns"] * len(missing_metabolites),
            "review_status": [""] * len(missing_metabolites),
            "ambiguous_flag": [False] * len(missing_metabolites),
            "notes": [""] * len(missing_metabolites),
        }
    )
    if resolved.empty:
        return seeded_rows
    return pd.concat([resolved, seeded_rows], ignore_index=True)


def _seed_global_profile_registry(registry: pd.DataFrame) -> pd.DataFrame:
    resolved = registry.copy()
    if resolved["model_id"].astype(str).str.strip().str.lower().eq(GLOBAL_PROFILE_MODEL_ID).any():
        return resolved

    global_profile = pd.DataFrame.from_records(
        [
            {
                "model_id": GLOBAL_PROFILE_MODEL_ID,
                "model_label": "Global Metabolite Profile",
                "model_tier": "primary",
                "model_status": "primary",
                "feature_kind": "continuous_abundance",
                "distance_kind": "correlation",
                "description": "All matrix metabolites",
                "authority": "user",
                "notes": "",
            }
        ]
    )
    return pd.concat([resolved, global_profile], ignore_index=True)


def _ensure_membership_columns(frame: pd.DataFrame) -> pd.DataFrame:
    normalized = frame.copy()
    for column in MODEL_MEMBERSHIP_OPTIONAL_COLUMNS:
        if column not in normalized.columns:
            normalized[column] = "" if column != "ambiguous_flag" else False

    normalized["membership_source"] = normalized["membership_source"].astype(str).str.strip()
    normalized["review_status"] = normalized["review_status"].astype(str).str.strip()
    normalized["notes"] = normalized["notes"].astype(str).str.strip()
    normalized["ambiguous_flag"] = normalized["ambiguous_flag"].astype(bool)
    return normalized[list(MODEL_MEMBERSHIP_REQUIRED_COLUMNS + MODEL_MEMBERSHIP_OPTIONAL_COLUMNS)]


def _validate_membership_against_registry(membership: pd.DataFrame, registry: pd.DataFrame) -> pd.DataFrame:
    if membership.empty:
        return membership

    registry_model_ids = set(registry["model_id"].astype(str))
    membership_model_ids = set(membership["model_id"].astype(str))
    unknown = sorted(membership_model_ids.difference(registry_model_ids))
    if unknown:
        raise ValueError(f"model_membership references unknown model_id values: {', '.join(unknown)}")

    return membership


def _validate_membership_against_matrix(membership: pd.DataFrame, matrix: pd.DataFrame) -> None:
    matrix_metabolites = set(matrix.columns.astype(str))
    membership_metabolites = set(membership["metabolite_name"].astype(str))
    unknown = sorted(membership_metabolites.difference(matrix_metabolites))
    if unknown:
        raise ValueError(f"model_membership metabolites must exist in the matrix: {', '.join(unknown)}")


def _get_model_registry_row(resolved_inputs: dict[str, pd.DataFrame], model_id: str) -> pd.Series:
    registry = resolved_inputs["model_registry_resolved"]
    matches = registry.loc[registry["model_id"].astype(str).str.strip().str.lower() == model_id]
    if matches.empty:
        raise ValueError(f"unknown model_id: {model_id}")
    if len(matches) != 1:
        raise ValueError(f"model_id must resolve to exactly one row: {model_id}")
    return matches.iloc[0]


def _select_model_input_matrix(resolved_inputs: dict[str, pd.DataFrame], model_id: str) -> pd.DataFrame:
    matrix = resolved_inputs["matrix"]
    mapping = resolved_inputs["stimulus_sample_map"]
    ordered_sample_ids = mapping["sample_id"].astype(str).tolist()
    ordered_stimuli = mapping["stimulus"].astype(str).tolist()
    selected_metabolites = _select_model_metabolites(resolved_inputs, model_id)
    ordered_matrix = matrix.loc[ordered_sample_ids, selected_metabolites].copy()
    ordered_matrix.index = pd.Index(ordered_stimuli, name="stimulus")
    ordered_matrix.columns = ordered_matrix.columns.astype(str)
    return ordered_matrix


def _select_model_metabolites(resolved_inputs: dict[str, pd.DataFrame], model_id: str) -> list[str]:
    matrix = resolved_inputs["matrix"]
    if model_id == GLOBAL_PROFILE_MODEL_ID:
        return matrix.columns.astype(str).tolist()

    membership = resolved_inputs["model_membership_resolved"]
    selected = membership.loc[
        membership["model_id"].astype(str).str.strip().str.lower() == model_id,
        "metabolite_name",
    ].astype(str)
    return selected.drop_duplicates().tolist()


def _drop_zero_variance_features(
    feature_matrix: pd.DataFrame,
    *,
    model_id: str,
    feature_kind: str,
    threshold: float | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if feature_matrix.shape[1] == 0:
        return feature_matrix.copy(), pd.DataFrame(columns=list(MODEL_FEATURE_QC_COLUMNS))

    variances = feature_matrix.var(axis=0, ddof=0).fillna(0.0)
    retained_mask = variances.gt(0.0)
    qc_rows: list[dict[str, object]] = []
    for metabolite_name in feature_matrix.columns.astype(str):
        retained = bool(retained_mask.loc[metabolite_name])
        qc_rows.append(
            {
                "model_id": model_id,
                "metabolite_name": metabolite_name,
                "feature_kind": feature_kind,
                "retained": retained,
                "status": "retained" if retained else "dropped",
                "filter_reason": "retained" if retained else "zero_variance",
                "threshold": threshold,
            }
        )

    filtered = feature_matrix.loc[:, retained_mask].copy()
    qc = pd.DataFrame.from_records(qc_rows, columns=list(MODEL_FEATURE_QC_COLUMNS))
    return filtered, qc


def _zscore_columns(feature_matrix: pd.DataFrame) -> pd.DataFrame:
    if feature_matrix.shape[1] == 0:
        return feature_matrix.copy()

    means = feature_matrix.mean(axis=0)
    stds = feature_matrix.std(axis=0, ddof=0).replace(0.0, 1.0)
    standardized = feature_matrix.subtract(means, axis="columns").divide(stds, axis="columns")
    standardized.index = feature_matrix.index
    standardized.columns = feature_matrix.columns
    return standardized


def _update_primary_ranking_exclusion(
    resolved_inputs: dict[str, pd.DataFrame],
    model_id: str,
    *,
    retained_feature_count: int,
) -> None:
    registry = resolved_inputs["model_registry_resolved"].copy()
    if "excluded_from_primary_ranking" not in registry.columns:
        registry["excluded_from_primary_ranking"] = False

    model_mask = registry["model_id"].astype(str).str.strip().str.lower() == model_id
    if not model_mask.any():
        raise ValueError(f"unknown model_id: {model_id}")

    is_primary = registry.loc[model_mask, "model_tier"].astype(str).str.strip().str.lower().eq(PRIMARY_TIER_VALUE)
    exclude = bool(is_primary.iloc[0] and retained_feature_count < PRIMARY_FEATURE_COUNT_MINIMUM)
    registry.loc[model_mask, "excluded_from_primary_ranking"] = exclude
    resolved_inputs["model_registry_resolved"] = registry


def _store_model_feature_qc(resolved_inputs: dict[str, pd.DataFrame], feature_qc: pd.DataFrame) -> None:
    existing = resolved_inputs.get("model_feature_qc")
    if existing is None or existing.empty:
        resolved_inputs["model_feature_qc"] = feature_qc.copy()
        return

    model_ids = feature_qc["model_id"].drop_duplicates().astype(str).tolist() if not feature_qc.empty else []
    remaining = existing.loc[~existing["model_id"].astype(str).isin(model_ids)].copy()
    resolved_inputs["model_feature_qc"] = pd.concat([remaining, feature_qc], ignore_index=True)


def _compute_model_distance_matrix(feature_matrix: pd.DataFrame, *, distance_kind: str) -> np.ndarray:
    values = feature_matrix.to_numpy(dtype=float, copy=False)
    if distance_kind == "correlation":
        return _pairwise_correlation_distance(values)
    if distance_kind == "euclidean":
        return _pairwise_euclidean_distance(values)
    if distance_kind == "jaccard":
        return _pairwise_jaccard_distance(values)
    raise ValueError(f"unsupported distance_kind: {distance_kind}")


def _validate_correlation_distance_inputs(feature_matrix: pd.DataFrame, *, model_id: str) -> None:
    if feature_matrix.shape[1] < 2:
        raise ValueError(
            f"correlation-based RDMs require at least 2 retained features for model {model_id}"
        )

    values = feature_matrix.to_numpy(dtype=float, copy=False)
    centered = values - values.mean(axis=1, keepdims=True)
    row_norms = np.linalg.norm(centered, axis=1)
    if np.any(row_norms == 0.0):
        raise ValueError(f"correlation-based RDMs require non-constant stimulus rows for model {model_id}")


def _pairwise_correlation_distance(values: np.ndarray) -> np.ndarray:
    n_rows = values.shape[0]
    if values.shape[1] == 0:
        return _empty_distance_matrix(n_rows)

    centered = values - values.mean(axis=1, keepdims=True)
    norms = np.linalg.norm(centered, axis=1)
    denominator = np.outer(norms, norms)
    similarity = np.divide(
        centered @ centered.T,
        denominator,
        out=np.zeros((n_rows, n_rows), dtype=float),
        where=denominator > 0,
    )
    np.fill_diagonal(similarity, 1.0)
    distance = 1.0 - similarity
    np.fill_diagonal(distance, 0.0)
    return distance


def _pairwise_euclidean_distance(values: np.ndarray) -> np.ndarray:
    n_rows = values.shape[0]
    if values.shape[1] == 0:
        return _empty_distance_matrix(n_rows)

    deltas = values[:, np.newaxis, :] - values[np.newaxis, :, :]
    distance = np.sqrt(np.sum(deltas * deltas, axis=2))
    np.fill_diagonal(distance, 0.0)
    return distance


def _pairwise_jaccard_distance(values: np.ndarray) -> np.ndarray:
    binary = values.astype(bool)
    intersections = np.logical_and(binary[:, np.newaxis, :], binary[np.newaxis, :, :]).sum(axis=2)
    unions = np.logical_or(binary[:, np.newaxis, :], binary[np.newaxis, :, :]).sum(axis=2)
    similarity = np.divide(
        intersections,
        unions,
        out=np.ones(intersections.shape, dtype=float),
        where=unions > 0,
    )
    distance = 1.0 - similarity
    np.fill_diagonal(distance, 0.0)
    return distance


def _empty_distance_matrix(n_rows: int) -> np.ndarray:
    distance = np.full((n_rows, n_rows), np.nan, dtype=float)
    np.fill_diagonal(distance, 0.0)
    return distance
