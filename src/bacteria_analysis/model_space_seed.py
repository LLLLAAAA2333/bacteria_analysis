"""Builder utilities for provisional Stage 3 model-space inputs."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
import hashlib
import json
from pathlib import Path
import re
from urllib import error, parse, request

import pandas as pd
from openpyxl import load_workbook

from bacteria_analysis.io import write_json
from bacteria_analysis.model_space import (
    METABOLITE_NAME_CANONICAL_OVERRIDES,
    build_stimulus_sample_map,
    load_model_registry,
    read_metabolite_matrix,
    resolve_model_inputs,
)

QUERY_CANDIDATE_DELIMITER = " | "
IDENTITY_EVIDENCE_COLUMNS = (
    "original_header",
    "normalized_name",
    "stage3_metabolite_name",
    "canonical_compound_name",
    "resolution_status",
    "source",
    "pubchem_cid",
    "chebi_id",
    "hmdb_id",
    "inchikey",
    "smiles",
    "synonyms",
    "match_score",
    "match_reason",
    "evidence_json",
)
TAXONOMY_ENRICHMENT_COLUMNS = (
    "normalized_name",
    "pubchem_cid",
    "chebi_id",
    "hmdb_id",
    "inchikey",
    "chebi_parent_terms",
    "hmdb_super_class",
    "hmdb_class",
    "hmdb_sub_class",
    "classyfire_kingdom",
    "classyfire_superclass",
    "classyfire_class",
    "classyfire_subclass",
    "pathway_tags",
    "annotation_confidence",
    "annotation_status",
    "annotation_evidence",
)
MODEL_MEMBERSHIP_COLUMNS = (
    "model_id",
    "metabolite_name",
    "membership_source",
    "review_status",
    "ambiguous_flag",
    "notes",
)
MEMBERSHIP_REVIEW_QUEUE_COLUMNS = (
    "model_id",
    "metabolite_name",
    "membership_source",
    "review_status",
    "ambiguous_flag",
    "reason",
    "matched_rule",
    "matched_field",
    "matched_value",
    "notes",
)
MEMBERSHIP_RULE_EVIDENCE_COLUMNS = (
    "model_id",
    "metabolite_name",
    "matched_field",
    "matched_value",
    "rule_name",
    "decision",
)
RULE_VERSION = "stage3-model-space-auto-seed-v1"
MODEL_RULE_KEYWORDS: dict[str, tuple[str, ...]] = {
    "bile_acid": ("bile acid", "bile acids", "cholic acid", "muricholic acid"),
    "indole_tryptophan": ("indole", "tryptophan", "kynurenine"),
    "phenyl_phenol": ("phenyl", "phenol", "phenolic", "benzoic acid"),
    "tca_organic_acid": ("citric acid cycle", "tricarboxylic acid", "succinate", "fumarate", "malate", "citrate"),
    "nucleotide_energy": ("nucleotide", "adenosine", "guanosine", "atp", "adp", "amp"),
    "broad_aromatic_union": ("aromatic", "indole", "phenyl", "phenol", "benzoic"),
}
PATHWAY_TAG_HINTS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("bile_acid", MODEL_RULE_KEYWORDS["bile_acid"]),
    ("indole_tryptophan", MODEL_RULE_KEYWORDS["indole_tryptophan"]),
    ("phenyl_phenol", MODEL_RULE_KEYWORDS["phenyl_phenol"]),
    ("tca_organic_acid", MODEL_RULE_KEYWORDS["tca_organic_acid"]),
    ("nucleotide_energy", MODEL_RULE_KEYWORDS["nucleotide_energy"]),
    ("broad_aromatic_union", MODEL_RULE_KEYWORDS["broad_aromatic_union"]),
)
GREEK_REPLACEMENTS = {
    "α": "alpha",
    "β": "beta",
    "γ": "gamma",
    "δ": "delta",
    "ω": "omega",
}
PRIME_CHARACTERS = {"′", "’", "‘", "ʼ", "ʹ", "ˈ", "´", "`"}
FINAL_PARENTHESES_RE = re.compile(r"\(([^()]+)\)\s*$")


@dataclass(frozen=True)
class RuleMatch:
    model_id: str
    matched_field: str
    matched_value: str
    keyword: str


def read_raw_metabolite_headers(matrix_path: str | Path) -> list[object]:
    """Read the first worksheet header row without canonicalizing metabolite names."""

    workbook = load_workbook(Path(matrix_path), read_only=True, data_only=True)
    try:
        first_sheet = workbook.worksheets[0]
        header_row = next(first_sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    finally:
        workbook.close()
    return list(header_row[1:])


def normalize_metabolite_header(header: object) -> dict[str, object]:
    """Normalize one raw matrix header into a deterministic lookup record."""

    original_header = header if isinstance(header, str) else ("" if header is None else str(header))
    trimmed = original_header.strip()
    if not trimmed:
        raise ValueError("metabolite headers must be non-empty")

    normalized_name, notes = _normalize_header_text(trimmed)
    alias = _extract_final_alias(normalized_name)
    query_candidates = _build_query_candidates(normalized_name, alias)
    return {
        "original_header": original_header,
        "normalized_name": normalized_name,
        "stage3_metabolite_name": _canonicalize_metabolite_name(normalized_name),
        "alias_in_parentheses": alias,
        "query_candidates": query_candidates,
        "normalization_notes": "; ".join(notes),
    }


def build_normalized_header_table(headers: list[object]) -> pd.DataFrame:
    """Normalize the full matrix header row and validate uniqueness."""

    records = [normalize_metabolite_header(header) for header in headers]
    frame = pd.DataFrame.from_records(
        records,
        columns=(
            "original_header",
            "normalized_name",
            "stage3_metabolite_name",
            "alias_in_parentheses",
            "query_candidates",
            "normalization_notes",
        ),
    )
    if frame["normalized_name"].duplicated().any():
        duplicates = frame.loc[frame["normalized_name"].duplicated(keep=False), "normalized_name"].tolist()
        raise ValueError(f"normalized metabolite names must be unique: {', '.join(dict.fromkeys(duplicates))}")
    if frame["stage3_metabolite_name"].duplicated().any():
        duplicates = frame.loc[
            frame["stage3_metabolite_name"].duplicated(keep=False),
            "stage3_metabolite_name",
        ].tolist()
        raise ValueError(f"canonical Stage 3 metabolite names must be unique: {', '.join(dict.fromkeys(duplicates))}")
    return frame


def load_identity_evidence_cache(path: str | Path | None) -> pd.DataFrame:
    """Load cached identity evidence or return an empty schema-valid table."""

    if path is None:
        return _empty_frame(IDENTITY_EVIDENCE_COLUMNS)

    frame = pd.read_csv(Path(path), dtype=str).fillna("")
    _require_columns(frame, IDENTITY_EVIDENCE_COLUMNS, "identity evidence cache")
    frame = frame.loc[:, list(IDENTITY_EVIDENCE_COLUMNS)].copy()
    frame["match_score"] = pd.to_numeric(frame["match_score"], errors="coerce").fillna(0.0)
    frame["normalized_name"] = frame["normalized_name"].astype(str).str.strip()
    if frame["normalized_name"].duplicated().any():
        raise ValueError("identity evidence cache normalized_name values must be unique")
    return frame


def resolve_identity_from_cached_candidates(
    normalized_record: dict[str, object],
    *,
    cached_payloads: list[dict[str, object]],
) -> dict[str, object]:
    """Resolve one normalized header from cached PubChem-style candidate payloads."""

    candidates = _collect_candidate_records(cached_payloads)
    if not candidates:
        return _default_identity_row(normalized_record, match_reason="no_identity_candidates")

    scored_candidates = [
        {
            **candidate,
            **_score_identity_candidate(normalized_record, candidate),
        }
        for candidate in candidates
    ]
    scored_candidates = [candidate for candidate in scored_candidates if float(candidate["match_score"]) > 0]
    if not scored_candidates:
        return _default_identity_row(normalized_record, match_reason="no_identity_candidates")

    scored_candidates.sort(key=lambda candidate: float(candidate["match_score"]), reverse=True)
    top_score = float(scored_candidates[0]["match_score"])
    top_candidates = [candidate for candidate in scored_candidates if float(candidate["match_score"]) == top_score]

    if len(top_candidates) > 1:
        evidence = {
            "query_candidates": list(normalized_record.get("query_candidates", [])),
            "top_candidates": [_compact_candidate(candidate) for candidate in top_candidates[:3]],
        }
        return {
            **_identity_base_row(normalized_record),
            "canonical_compound_name": "",
            "resolution_status": "multi_hit_needs_review",
            "source": "pubchem",
            "pubchem_cid": "",
            "chebi_id": "",
            "hmdb_id": "",
            "inchikey": "",
            "smiles": "",
            "synonyms": "",
            "match_score": top_score,
            "match_reason": "multiple_equal_candidates",
            "evidence_json": json.dumps(evidence, sort_keys=True, ensure_ascii=False),
        }

    best = top_candidates[0]
    has_inchikey = bool(str(best.get("inchikey", "")).strip())
    if best["match_kind"] in {"exact_candidate_name_match", "exact_synonym_match"} and has_inchikey:
        resolution_status = "resolved_high_confidence"
    else:
        resolution_status = "resolved_low_confidence"

    evidence = {
        "query_candidates": list(normalized_record.get("query_candidates", [])),
        "selected_candidate": _compact_candidate(best),
    }
    return {
        **_identity_base_row(normalized_record),
        "canonical_compound_name": str(best.get("candidate_name", "")).strip(),
        "resolution_status": resolution_status,
        "source": str(best.get("source", "pubchem")).strip(),
        "pubchem_cid": str(best.get("pubchem_cid", "")).strip(),
        "chebi_id": str(best.get("chebi_id", "")).strip(),
        "hmdb_id": str(best.get("hmdb_id", "")).strip(),
        "inchikey": str(best.get("inchikey", "")).strip(),
        "smiles": str(best.get("smiles", "")).strip(),
        "synonyms": _serialize_list(best.get("synonyms", [])),
        "match_score": float(best["match_score"]),
        "match_reason": str(best["match_kind"]),
        "evidence_json": json.dumps(evidence, sort_keys=True, ensure_ascii=False),
    }


def fetch_pubchem_payload(query: str, *, timeout_seconds: float = 20.0) -> dict[str, object]:
    """Fetch a normalized PubChem candidate bundle for one query."""

    query = str(query).strip()
    if not query:
        raise ValueError("query must be non-empty")

    quoted_query = parse.quote(query)
    cid_payload = _read_json_url(
        f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{quoted_query}/cids/JSON",
        timeout_seconds=timeout_seconds,
        allow_not_found=True,
    )
    cids = cid_payload.get("IdentifierList", {}).get("CID", [])
    candidates: list[dict[str, object]] = []
    for cid in cids[:5]:
        property_payload = _read_json_url(
            (
                "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/"
                f"{cid}/property/Title,CanonicalSMILES,InChIKey/JSON"
            ),
            timeout_seconds=timeout_seconds,
            allow_not_found=True,
        )
        synonym_payload = _read_json_url(
            f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/synonyms/JSON",
            timeout_seconds=timeout_seconds,
            allow_not_found=True,
        )
        property_rows = property_payload.get("PropertyTable", {}).get("Properties", [])
        property_row = property_rows[0] if property_rows else {}
        synonym_rows = synonym_payload.get("InformationList", {}).get("Information", [])
        synonym_row = synonym_rows[0] if synonym_rows else {}
        candidates.append(
            {
                "source": "pubchem",
                "query": query,
                "pubchem_cid": str(cid),
                "candidate_name": str(property_row.get("Title", "")).strip(),
                "synonyms": synonym_row.get("Synonym", []) or [],
                "smiles": str(property_row.get("CanonicalSMILES", "")).strip(),
                "inchikey": str(property_row.get("InChIKey", "")).strip(),
            }
        )

    return {"query": query, "source": "pubchem", "candidates": candidates}


def load_taxonomy_enrichment_cache(path: str | Path | None) -> pd.DataFrame:
    """Load cached taxonomy enrichment or return an empty schema-valid table."""

    if path is None:
        return _empty_frame(TAXONOMY_ENRICHMENT_COLUMNS)

    frame = pd.read_csv(Path(path), dtype=str).fillna("")
    _require_columns(frame, TAXONOMY_ENRICHMENT_COLUMNS, "taxonomy enrichment cache")
    frame = frame.loc[:, list(TAXONOMY_ENRICHMENT_COLUMNS)].copy()
    frame["normalized_name"] = frame["normalized_name"].astype(str).str.strip()
    if frame["normalized_name"].duplicated().any():
        raise ValueError("taxonomy enrichment normalized_name values must be unique")
    return frame


def merge_identity_and_taxonomy_evidence(
    identity_evidence: pd.DataFrame,
    taxonomy_enrichment: pd.DataFrame,
) -> pd.DataFrame:
    """Join aligned identity evidence with optional taxonomy enrichment."""

    merged = identity_evidence.copy()
    if taxonomy_enrichment.empty:
        for column in TAXONOMY_ENRICHMENT_COLUMNS:
            if column not in merged.columns:
                merged[column] = ""
        return merged

    taxonomy = taxonomy_enrichment.copy()
    joined = merged.merge(taxonomy, on="normalized_name", how="left", suffixes=("", "_taxonomy"))
    _validate_taxonomy_identity_alignment(joined)
    for column in ("pubchem_cid", "chebi_id", "hmdb_id", "inchikey"):
        taxonomy_column = f"{column}_taxonomy"
        joined[column] = joined[column].where(
            joined[column].astype(str).str.strip() != "",
            joined[taxonomy_column].fillna(""),
        )
        joined = joined.drop(columns=[taxonomy_column])
    return joined.fillna("")


def build_metabolite_annotation_from_evidence(
    normalized_headers: pd.DataFrame,
    identity_evidence: pd.DataFrame,
) -> pd.DataFrame:
    """Project rich evidence into the simplified Stage 3 annotation CSV."""

    merged = normalized_headers.loc[:, ["normalized_name", "stage3_metabolite_name"]].merge(
        identity_evidence,
        on="normalized_name",
        how="left",
        suffixes=("", "_evidence"),
    )
    rows: list[dict[str, object]] = []
    for _, row in merged.iterrows():
        resolution_status = str(row.get("resolution_status", "")).strip()
        if resolution_status == "resolved_high_confidence":
            review_status = "auto_high_confidence"
            ambiguous_flag = False
        elif resolution_status == "unresolved":
            review_status = "unresolved"
            ambiguous_flag = True
        else:
            review_status = "needs_review"
            ambiguous_flag = True

        rows.append(
            {
                "metabolite_name": str(row["stage3_metabolite_name"]),
                "superclass": _first_non_empty(
                    row.get("hmdb_super_class", ""),
                    row.get("classyfire_superclass", ""),
                    row.get("classyfire_kingdom", ""),
                ),
                "subclass": _first_non_empty(
                    row.get("hmdb_sub_class", ""),
                    row.get("hmdb_class", ""),
                    row.get("classyfire_subclass", ""),
                    row.get("classyfire_class", ""),
                    row.get("chebi_parent_terms", ""),
                ),
                "pathway_tag": _derive_pathway_tag(row),
                "annotation_source": "auto_identity_resolution"
                if resolution_status != "unresolved"
                else "unresolved_identity",
                "review_status": review_status,
                "ambiguous_flag": ambiguous_flag,
                "notes": _annotation_notes(row),
            }
        )

    return pd.DataFrame.from_records(
        rows,
        columns=(
            "metabolite_name",
            "superclass",
            "subclass",
            "pathway_tag",
            "annotation_source",
            "review_status",
            "ambiguous_flag",
            "notes",
        ),
    )


def build_model_membership_from_rules(
    annotation: pd.DataFrame,
    registry: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Generate conservative model membership and review queue rows from annotation."""

    active_registry = registry.loc[registry["model_status"].astype(str).str.strip().str.lower() != "excluded"].copy()
    active_registry["model_id"] = active_registry["model_id"].astype(str).str.strip().str.lower()
    active_registry = active_registry.loc[active_registry["model_id"] != "global_profile"].copy()
    _validate_registry_rule_coverage(active_registry)

    membership_rows: list[dict[str, object]] = []
    review_rows: list[dict[str, object]] = []
    evidence_rows: list[dict[str, object]] = []
    seen_memberships: set[tuple[str, str]] = set()
    seen_reviews: set[tuple[str, str, str, str]] = set()

    for _, row in annotation.iterrows():
        metabolite_name = str(row["metabolite_name"]).strip()
        review_status = str(row.get("review_status", "")).strip()
        matches = _collect_rule_matches(row, active_registry)

        if not matches and review_status in {"needs_review", "unresolved"}:
            review_row = {
                "model_id": "",
                "metabolite_name": metabolite_name,
                "membership_source": "unresolved_identity" if review_status == "unresolved" else "auto_taxonomy_rule",
                "review_status": review_status,
                "ambiguous_flag": True,
                "reason": "unresolved_identity" if review_status == "unresolved" else "needs_review_without_rule_match",
                "matched_rule": "",
                "matched_field": "",
                "matched_value": "",
                "notes": str(row.get("notes", "")).strip(),
            }
            review_key = ("", metabolite_name, review_row["review_status"], review_row["reason"])
            if review_key not in seen_reviews:
                review_rows.append(review_row)
                seen_reviews.add(review_key)
            continue

        for match in matches:
            evidence_rows.append(
                {
                    "model_id": match.model_id,
                    "metabolite_name": metabolite_name,
                    "matched_field": match.matched_field,
                    "matched_value": match.matched_value,
                    "rule_name": match.model_id,
                    "decision": "include_auto_high_confidence"
                    if review_status == "auto_high_confidence"
                    else "queue_for_review",
                }
            )
            if review_status == "auto_high_confidence":
                membership_key = (match.model_id, metabolite_name)
                if membership_key in seen_memberships:
                    continue
                membership_rows.append(
                    {
                        "model_id": match.model_id,
                        "metabolite_name": metabolite_name,
                        "membership_source": "auto_taxonomy_rule",
                        "review_status": "auto_high_confidence",
                        "ambiguous_flag": False,
                        "notes": f"rule={match.model_id}",
                    }
                )
                seen_memberships.add(membership_key)
                continue

            review_key = (match.model_id, metabolite_name, review_status, match.matched_field)
            if review_key in seen_reviews:
                continue
            review_rows.append(
                {
                    "model_id": match.model_id,
                    "metabolite_name": metabolite_name,
                    "membership_source": "unresolved_identity" if review_status == "unresolved" else "auto_taxonomy_rule",
                    "review_status": "unresolved" if review_status == "unresolved" else "needs_review",
                    "ambiguous_flag": True,
                    "reason": "unresolved_identity" if review_status == "unresolved" else "needs_review",
                    "matched_rule": match.model_id,
                    "matched_field": match.matched_field,
                    "matched_value": match.matched_value,
                    "notes": str(row.get("notes", "")).strip(),
                }
            )
            seen_reviews.add(review_key)

    membership = pd.DataFrame.from_records(membership_rows, columns=MODEL_MEMBERSHIP_COLUMNS)
    review_queue = pd.DataFrame.from_records(review_rows, columns=MEMBERSHIP_REVIEW_QUEUE_COLUMNS)
    evidence = pd.DataFrame.from_records(evidence_rows, columns=MEMBERSHIP_RULE_EVIDENCE_COLUMNS)
    return membership, review_queue, evidence


def build_model_space_manifest(
    *,
    matrix_path: str | Path,
    preprocess_root: str | Path,
    trial_metadata_path: str | Path,
    registry_path: str | Path,
    output_root: str | Path,
    identity_evidence_path: str | Path | None,
    taxonomy_enrichment_path: str | Path | None,
    cache_version: str,
    refresh_pubchem_cache: bool,
) -> dict[str, object]:
    """Build a reproducibility manifest for generated model-space outputs."""

    matrix_path = Path(matrix_path)
    preprocess_root = Path(preprocess_root)
    trial_metadata_path = Path(trial_metadata_path)
    registry_path = Path(registry_path)
    output_root = Path(output_root)
    output_files = {
        "stimulus_sample_map": "stimulus_sample_map.csv",
        "metabolite_annotation": "metabolite_annotation.csv",
        "model_registry": "model_registry.csv",
        "model_membership": "model_membership.csv",
        "model_membership_review_queue": "model_membership_review_queue.csv",
        "model_space_manifest": "model_space_manifest.json",
        "cache_normalized_headers": "cache/normalized_headers.csv",
        "cache_identity_resolution_evidence": "cache/identity_resolution_evidence.csv",
        "cache_taxonomy_enrichment_evidence": "cache/taxonomy_enrichment_evidence.csv",
        "cache_membership_rule_evidence": "cache/membership_rule_evidence.csv",
    }

    return {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "rule_version": RULE_VERSION,
        "cache_version": cache_version,
        "matrix_path": str(matrix_path),
        "matrix_sha256": _sha256_file(matrix_path),
        "preprocess_root": str(preprocess_root),
        "trial_metadata_path": str(trial_metadata_path),
        "trial_metadata_sha256": _sha256_file(trial_metadata_path),
        "registry_path": str(registry_path),
        "registry_sha256": _sha256_file(registry_path),
        "identity_evidence_cache_path": None if identity_evidence_path is None else str(Path(identity_evidence_path)),
        "identity_evidence_cache_sha256": None
        if identity_evidence_path is None
        else _sha256_file(Path(identity_evidence_path)),
        "taxonomy_enrichment_cache_path": None
        if taxonomy_enrichment_path is None
        else str(Path(taxonomy_enrichment_path)),
        "taxonomy_enrichment_cache_sha256": None
        if taxonomy_enrichment_path is None
        else _sha256_file(Path(taxonomy_enrichment_path)),
        "pubchem_mode": "live_refresh" if refresh_pubchem_cache else "cache_only",
        "chebi_mode": "cache_only",
        "hmdb_mode": "cache_only",
        "classyfire_mode": "cache_only",
        "output_root": str(output_root),
        "outputs": output_files,
    }


def build_model_space(
    *,
    matrix_path: str | Path,
    preprocess_root: str | Path,
    registry_path: str | Path,
    output_root: str | Path,
    identity_evidence_path: str | Path | None = None,
    taxonomy_enrichment_path: str | Path | None = None,
    cache_version: str = "manual-cache-v1",
    refresh_pubchem_cache: bool = False,
) -> dict[str, Path]:
    """Build a provisional Stage 3 model-space directory."""

    matrix_path = Path(matrix_path)
    preprocess_root = Path(preprocess_root)
    registry_path = Path(registry_path)
    output_root = Path(output_root)
    cache_dir = output_root / "cache"
    trial_metadata_path = preprocess_root / "trial_level" / "trial_metadata.parquet"

    matrix = read_metabolite_matrix(matrix_path)
    raw_headers = read_raw_metabolite_headers(matrix_path)
    if len(raw_headers) != matrix.shape[1]:
        raise ValueError("raw matrix header count must match canonical Stage 3 metabolite count")
    normalized_headers = build_normalized_header_table(raw_headers)
    normalized_headers["stage3_metabolite_name"] = matrix.columns.astype(str).tolist()
    normalized_headers["query_candidates"] = [
        _unique_non_empty([stage3_name, *query_candidates])
        for stage3_name, query_candidates in zip(
            normalized_headers["stage3_metabolite_name"].tolist(),
            normalized_headers["query_candidates"].tolist(),
            strict=True,
        )
    ]

    trial_metadata = pd.read_parquet(trial_metadata_path)
    stimulus_sample_map = build_stimulus_sample_map(trial_metadata, matrix_sample_ids=matrix.index)
    model_registry = load_model_registry(registry_path)

    identity_evidence = load_identity_evidence_cache(identity_evidence_path)
    identity_evidence = _align_identity_evidence(
        normalized_headers,
        identity_evidence=identity_evidence,
        refresh_pubchem_cache=refresh_pubchem_cache,
    )
    taxonomy_enrichment = load_taxonomy_enrichment_cache(taxonomy_enrichment_path)
    merged_evidence = merge_identity_and_taxonomy_evidence(identity_evidence, taxonomy_enrichment)
    metabolite_annotation = build_metabolite_annotation_from_evidence(normalized_headers, merged_evidence)
    model_membership, review_queue, membership_rule_evidence = build_model_membership_from_rules(
        metabolite_annotation,
        model_registry,
    )

    output_root.mkdir(parents=True, exist_ok=True)
    cache_dir.mkdir(parents=True, exist_ok=True)
    normalized_headers_for_write = normalized_headers.copy()
    normalized_headers_for_write["query_candidates"] = normalized_headers_for_write["query_candidates"].map(
        _serialize_list
    )

    stimulus_sample_map_path = output_root / "stimulus_sample_map.csv"
    metabolite_annotation_path = output_root / "metabolite_annotation.csv"
    model_registry_path = output_root / "model_registry.csv"
    model_membership_path = output_root / "model_membership.csv"
    review_queue_path = output_root / "model_membership_review_queue.csv"
    normalized_headers_path = cache_dir / "normalized_headers.csv"
    identity_evidence_output_path = cache_dir / "identity_resolution_evidence.csv"
    taxonomy_evidence_output_path = cache_dir / "taxonomy_enrichment_evidence.csv"
    membership_rule_output_path = cache_dir / "membership_rule_evidence.csv"

    stimulus_sample_map.to_csv(stimulus_sample_map_path, index=False)
    metabolite_annotation.to_csv(metabolite_annotation_path, index=False)
    model_registry.to_csv(model_registry_path, index=False)
    model_membership.to_csv(model_membership_path, index=False)
    review_queue.to_csv(review_queue_path, index=False)
    normalized_headers_for_write.to_csv(normalized_headers_path, index=False)
    identity_evidence.to_csv(identity_evidence_output_path, index=False)
    merged_evidence.to_csv(taxonomy_evidence_output_path, index=False)
    membership_rule_evidence.to_csv(membership_rule_output_path, index=False)

    manifest = build_model_space_manifest(
        matrix_path=matrix_path,
        preprocess_root=preprocess_root,
        trial_metadata_path=trial_metadata_path,
        registry_path=registry_path,
        output_root=output_root,
        identity_evidence_path=identity_evidence_path,
        taxonomy_enrichment_path=taxonomy_enrichment_path,
        cache_version=cache_version,
        refresh_pubchem_cache=refresh_pubchem_cache,
    )
    manifest_path = output_root / "model_space_manifest.json"
    write_json(manifest, manifest_path)

    resolve_model_inputs(output_root, matrix_path)
    return {
        "output_root": output_root,
        "stimulus_sample_map": stimulus_sample_map_path,
        "metabolite_annotation": metabolite_annotation_path,
        "model_registry": model_registry_path,
        "model_membership": model_membership_path,
        "model_membership_review_queue": review_queue_path,
        "model_space_manifest": manifest_path,
        "cache_dir": cache_dir,
    }


def _normalize_header_text(value: str) -> tuple[str, list[str]]:
    text = value
    notes: list[str] = []
    for original, replacement in GREEK_REPLACEMENTS.items():
        if original in text:
            text = text.replace(original, replacement)
            notes.append(f"greek:{original}->{replacement}")
    for original in PRIME_CHARACTERS:
        if original in text:
            text = text.replace(original, "'")
            notes.append("prime:normalized")
    text = re.sub(r"\s+", " ", text).strip()
    return text, notes


def _extract_final_alias(value: str) -> str:
    match = FINAL_PARENTHESES_RE.search(value)
    if match is None:
        return ""
    return match.group(1).strip()


def _build_query_candidates(normalized_name: str, alias: str) -> list[str]:
    candidates = [normalized_name]
    aliasless = FINAL_PARENTHESES_RE.sub("", normalized_name).strip()
    if aliasless:
        candidates.append(aliasless)
    if alias:
        candidates.append(alias)
    return _unique_non_empty(candidates)


def _collect_candidate_records(cached_payloads: list[dict[str, object]]) -> list[dict[str, object]]:
    candidates: list[dict[str, object]] = []
    for payload in cached_payloads:
        if "candidates" in payload:
            source = str(payload.get("source", "pubchem")).strip() or "pubchem"
            for candidate in payload.get("candidates", []):
                record = dict(candidate)
                record.setdefault("source", source)
                candidates.append(record)
        else:
            record = dict(payload)
            record.setdefault("source", "pubchem")
            candidates.append(record)
    return candidates


def _score_identity_candidate(normalized_record: dict[str, object], candidate: dict[str, object]) -> dict[str, object]:
    query_candidates = [_normalize_match_text(value) for value in normalized_record.get("query_candidates", []) if str(value).strip()]
    candidate_name = _normalize_match_text(candidate.get("candidate_name", ""))
    synonyms = [_normalize_match_text(value) for value in candidate.get("synonyms", []) if str(value).strip()]
    if candidate_name in query_candidates:
        return {"match_score": 100.0, "match_kind": "exact_candidate_name_match"}
    if any(synonym in query_candidates for synonym in synonyms):
        return {"match_score": 100.0, "match_kind": "exact_synonym_match"}

    alias = _normalize_match_text(normalized_record.get("alias_in_parentheses", ""))
    aliasless = _normalize_match_text(normalized_record.get("normalized_name", ""))
    aliasless = FINAL_PARENTHESES_RE.sub("", aliasless).strip()
    if aliasless and candidate_name and (aliasless in candidate_name or candidate_name in aliasless):
        return {"match_score": 60.0, "match_kind": "loose_name_match"}
    if alias and (
        (candidate_name and alias in candidate_name)
        or any(alias in synonym or synonym in alias for synonym in synonyms)
    ):
        return {"match_score": 55.0, "match_kind": "loose_alias_match"}
    return {"match_score": 0.0, "match_kind": "no_match"}


def _identity_base_row(normalized_record: dict[str, object]) -> dict[str, object]:
    return {
        "original_header": str(normalized_record.get("original_header", "")),
        "normalized_name": str(normalized_record.get("normalized_name", "")).strip(),
        "stage3_metabolite_name": str(normalized_record.get("stage3_metabolite_name", "")).strip(),
    }


def _default_identity_row(normalized_record: dict[str, object], *, match_reason: str) -> dict[str, object]:
    return {
        **_identity_base_row(normalized_record),
        "canonical_compound_name": "",
        "resolution_status": "unresolved",
        "source": "",
        "pubchem_cid": "",
        "chebi_id": "",
        "hmdb_id": "",
        "inchikey": "",
        "smiles": "",
        "synonyms": "",
        "match_score": 0.0,
        "match_reason": match_reason,
        "evidence_json": "{}",
    }


def _compact_candidate(candidate: dict[str, object]) -> dict[str, object]:
    return {
        "pubchem_cid": str(candidate.get("pubchem_cid", "")).strip(),
        "candidate_name": str(candidate.get("candidate_name", "")).strip(),
        "inchikey": str(candidate.get("inchikey", "")).strip(),
        "match_kind": str(candidate.get("match_kind", "")).strip(),
        "match_score": float(candidate.get("match_score", 0.0)),
    }


def _align_identity_evidence(
    normalized_headers: pd.DataFrame,
    *,
    identity_evidence: pd.DataFrame,
    refresh_pubchem_cache: bool,
) -> pd.DataFrame:
    if identity_evidence.empty and not refresh_pubchem_cache:
        rows = [_default_identity_row(row.to_dict(), match_reason="no_identity_candidates") for _, row in normalized_headers.iterrows()]
        return pd.DataFrame.from_records(rows, columns=IDENTITY_EVIDENCE_COLUMNS)

    evidence_by_name = {
        str(row["normalized_name"]).strip(): row.to_dict()
        for _, row in identity_evidence.iterrows()
    }
    rows: list[dict[str, object]] = []
    for _, header_row in normalized_headers.iterrows():
        header_record = header_row.to_dict()
        normalized_name = str(header_record["normalized_name"]).strip()
        if normalized_name in evidence_by_name and not refresh_pubchem_cache:
            row = evidence_by_name[normalized_name]
            row["original_header"] = str(header_record["original_header"])
            row["stage3_metabolite_name"] = str(header_record["stage3_metabolite_name"])
            rows.append({column: row.get(column, "") for column in IDENTITY_EVIDENCE_COLUMNS})
            continue
        if refresh_pubchem_cache:
            payloads = [fetch_pubchem_payload(query) for query in header_record.get("query_candidates", [])]
            rows.append(resolve_identity_from_cached_candidates(header_record, cached_payloads=payloads))
        else:
            rows.append(_default_identity_row(header_record, match_reason="no_identity_candidates"))

    frame = pd.DataFrame.from_records(rows, columns=IDENTITY_EVIDENCE_COLUMNS)
    frame["match_score"] = pd.to_numeric(frame["match_score"], errors="coerce").fillna(0.0)
    return frame


def _derive_pathway_tag(row: pd.Series) -> str:
    raw_pathway_tags = str(row.get("pathway_tags", "")).strip()
    texts = [
        raw_pathway_tags,
        str(row.get("chebi_parent_terms", "")).strip(),
        str(row.get("hmdb_super_class", "")).strip(),
        str(row.get("hmdb_class", "")).strip(),
        str(row.get("hmdb_sub_class", "")).strip(),
        str(row.get("classyfire_superclass", "")).strip(),
        str(row.get("classyfire_class", "")).strip(),
        str(row.get("classyfire_subclass", "")).strip(),
        str(row.get("canonical_compound_name", "")).strip(),
        str(row.get("stage3_metabolite_name", "")).strip(),
    ]
    for pathway_tag, keywords in PATHWAY_TAG_HINTS:
        if any(_keyword_matches(keyword, text) for keyword in keywords for text in texts if text):
            return pathway_tag
    return raw_pathway_tags


def _annotation_notes(row: pd.Series) -> str:
    parts: list[str] = []
    canonical_name = str(row.get("canonical_compound_name", "")).strip()
    match_reason = str(row.get("match_reason", "")).strip()
    if canonical_name:
        parts.append(f"compound={canonical_name}")
    if match_reason:
        parts.append(f"reason={match_reason}")
    return "; ".join(parts)


def _validate_registry_rule_coverage(registry: pd.DataFrame) -> None:
    missing = sorted(set(registry["model_id"].astype(str)).difference(MODEL_RULE_KEYWORDS))
    if missing:
        raise ValueError(f"missing membership rules for model_id values: {', '.join(missing)}")


def _validate_taxonomy_identity_alignment(joined: pd.DataFrame) -> None:
    conflicts: list[str] = []
    for _, row in joined.iterrows():
        for column in ("pubchem_cid", "chebi_id", "hmdb_id", "inchikey"):
            identity_value = str(row.get(column, "")).strip()
            taxonomy_value = str(row.get(f"{column}_taxonomy", "")).strip()
            if identity_value and taxonomy_value and identity_value != taxonomy_value:
                conflicts.append(str(row.get("normalized_name", "")).strip())
                break
    if conflicts:
        raise ValueError(
            "taxonomy enrichment conflicts with identity evidence for normalized_name values: "
            + ", ".join(dict.fromkeys(conflicts))
        )


def _collect_rule_matches(annotation_row: pd.Series, registry: pd.DataFrame) -> list[RuleMatch]:
    matches: list[RuleMatch] = []
    searchable_fields = {
        "metabolite_name": str(annotation_row.get("metabolite_name", "")).strip(),
        "superclass": str(annotation_row.get("superclass", "")).strip(),
        "subclass": str(annotation_row.get("subclass", "")).strip(),
        "pathway_tag": str(annotation_row.get("pathway_tag", "")).strip(),
        "notes": str(annotation_row.get("notes", "")).strip(),
    }
    available_model_ids = set(registry["model_id"].astype(str).tolist())
    for model_id, keywords in MODEL_RULE_KEYWORDS.items():
        if model_id not in available_model_ids:
            continue
        for field_name, field_value in searchable_fields.items():
            for keyword in keywords:
                if _keyword_matches(keyword, field_value):
                    matches.append(
                        RuleMatch(
                            model_id=model_id,
                            matched_field=field_name,
                            matched_value=field_value,
                            keyword=keyword,
                        )
                    )
                    break
            else:
                continue
            break
    return matches


def _canonicalize_metabolite_name(value: object) -> str:
    if value is None:
        return ""
    normalized = str(value).strip()
    return METABOLITE_NAME_CANONICAL_OVERRIDES.get(normalized, normalized)


def _first_non_empty(*values: object) -> str:
    for value in values:
        text = str(value).strip()
        if text:
            return text
    return ""


def _normalize_match_text(value: object) -> str:
    return _normalize_header_text(str(value).strip())[0].lower()


def _keyword_matches(keyword: str, field_value: str) -> bool:
    normalized_value = _normalize_match_text(field_value).replace("_", " ")
    normalized_keyword = _normalize_match_text(keyword).replace("_", " ")
    if " " in normalized_keyword:
        return normalized_keyword in normalized_value
    tokens = re.findall(r"[a-z0-9']+", normalized_value)
    return normalized_keyword in tokens


def _serialize_list(value: object) -> str:
    if isinstance(value, list):
        return QUERY_CANDIDATE_DELIMITER.join(str(item).strip() for item in value if str(item).strip())
    return str(value).strip()


def _unique_non_empty(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        cleaned = str(value).strip()
        if not cleaned or cleaned in seen:
            continue
        result.append(cleaned)
        seen.add(cleaned)
    return result


def _empty_frame(columns: tuple[str, ...]) -> pd.DataFrame:
    return pd.DataFrame(columns=list(columns))


def _require_columns(frame: pd.DataFrame, required_columns: tuple[str, ...], label: str) -> None:
    missing = [column for column in required_columns if column not in frame.columns]
    if missing:
        raise ValueError(f"{label} must include columns: {', '.join(missing)}")


def _read_json_url(url: str, *, timeout_seconds: float, allow_not_found: bool) -> dict[str, object]:
    try:
        with request.urlopen(url, timeout=timeout_seconds) as response:
            return json.loads(response.read().decode("utf-8"))
    except error.HTTPError as exc:
        if allow_not_found and exc.code == 404:
            return {}
        raise


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
